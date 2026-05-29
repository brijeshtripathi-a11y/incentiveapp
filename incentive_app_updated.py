"""
Tele Annual (TA) Incentive Calculator - Apr 2026
All levels: L1 Exec, L2 Rel'n Mgr, L3 BM, L4 CH + Nursery
For Tele Annual CSD and Tele Annual KCD - SEPARATE from regular CSD/KCD
"""
import streamlit as st
import pandas as pd
import io
import io as _io
from datetime import date

st.set_page_config(page_title="TA Incentive Calculator", layout="wide", page_icon="📊")

# These dates are overridden at runtime by the month selector in the UI
CALC_DATE  = date(2026, 5, 31)   # Updated to May 2026
FY_START   = date(2026, 4, 1)    # Financial year start (Apr 1, fixed)

# BD→TA movement employees who qualify for "60D" nursery scheme
# Source: "BD to TA Movement Month" column in sir's TA sheet
_60D_NURSERY_EIDS = {"60504","117268","104639","112667","119273",
                     "106325","112044","107101","116951"}
EXCEL_EPOCH= date(1899, 12, 30)

IM_STAR_LEADER_KW = {"IM STAR","IM LEADER","CITY STAR","CITY LEADER","PREFERRED STAR",
    "PREFERRED LEADER","IM STAR PRO","IM LEADER PRO","PREFERRED STAR PRO",
    "PREFERRED LEADER PRO","IM STAR-ADDON","IM STAR PRO-ADDON",
    "PREFERRED STAR PRO-ADDON","PREFERRED LEADER-ADDON"}
IM_STAR_PRO_KW = {"IM STAR PRO","IM LEADER PRO","PREFERRED STAR PRO",
    "PREFERRED LEADER PRO","PREFERRED STAR PRO-ADDON","PREFERRED LEADER-ADDON"}
SS_PLUS_KW = {"IM STAR","IM LEADER","CITY STAR","CITY LEADER","PREFERRED STAR",
    "PREFERRED LEADER","IM STAR PRO","IM LEADER PRO","PREFERRED STAR PRO","PREFERRED LEADER PRO"}

# ──────────────────────────────────────────────────────────
# SLAB CONFIG (defaults + loader)
# ──────────────────────────────────────────────────────────
def build_ta_slab_config():
    csd_milestones = pd.DataFrame([
        {"Min_Ach_Pct":140,"Grid":8000},
        {"Min_Ach_Pct":120,"Grid":6000},
        {"Min_Ach_Pct":100,"Grid":4000},
    ])
    csd_targets = pd.DataFrame([
        {"Vintage":"0-90","Target_PCDV":1300},
        
        {"Vintage":"90-270","Target_PCDV":1800},
        {"Vintage":"270+","Target_PCDV":2000},
    ])
    csd_cmr_mult = pd.DataFrame([
        {"Min_CMR_Ach_Pct":120,"Multiplier":1.2},
        {"Min_CMR_Ach_Pct":100,"Multiplier":1.0},
        {"Min_CMR_Ach_Pct":0,"Multiplier":0.0},
    ])
    csd_params = pd.DataFrame([
        {"Parameter":"Incr_Rate_Pct","Value":3.0},
        {"Parameter":"CMR_Target_Pct","Value":40.0},
    ])
    csd_spot_2_6  = pd.DataFrame([{"Txn":3,"Inc":2250},{"Txn":2,"Inc":1500},{"Txn":1,"Inc":500}])
    csd_spot_7_12 = pd.DataFrame([{"Txn":3,"Inc":2250},{"Txn":2,"Inc":1500},{"Txn":1,"Inc":500}])
    csd_spot_20_30= pd.DataFrame([{"Txn":6,"Inc":4250},{"Txn":5,"Inc":3500},{"Txn":4,"Inc":1375},{"Txn":3,"Inc":1000}])

    kcd_milestones= csd_milestones.copy()
    kcd_targets = pd.DataFrame([
        {"Vintage":"0-270","Group":"KCD","Target_PCDV":7100},   # Confirmed from sir
        {"Vintage":"270-2Yr","Group":"KCD","Target_PCDV":8100},
        {"Vintage":"2Yr+","Group":"KCD","Target_PCDV":9100},
        {"Vintage":"0-270","Group":"KCD-25cr","Target_PCDV":6480},
        {"Vintage":"270-2Yr","Group":"KCD-25cr","Target_PCDV":9100},
        {"Vintage":"2Yr+","Group":"KCD-25cr","Target_PCDV":10000},
    ])
    kcd_spot_1_12 = pd.DataFrame([
        {"PCDV":9000,"Inc":8000},{"PCDV":7000,"Inc":5500},
        {"PCDV":6000,"Inc":4500},{"PCDV":5500,"Inc":4000},{"PCDV":4500,"Inc":3000},
    ])
    kcd_spot_20_30= kcd_spot_1_12.copy()
    kcd_25cr_1_12 = pd.DataFrame([
        {"PCDV":12000,"Inc":10000},{"PCDV":10000,"Inc":8000},
        {"PCDV":8000,"Inc":6000},{"PCDV":6000,"Inc":4000},
    ])
    kcd_25cr_20_30= kcd_25cr_1_12.copy()
    kcd_rm_1_12   = kcd_spot_1_12.copy()
    kcd_rm_20_30  = kcd_spot_20_30.copy()
    kcd_ss_mult   = pd.DataFrame([{"Min_SS_CMR_Pct":50,"Multiplier":1.25},{"Min_SS_CMR_Pct":0,"Multiplier":0.5}])
    kcd_params    = pd.DataFrame([{"Parameter":"Min_SS_Sent","Value":3}])

    nursery_params= pd.DataFrame([
        {"Parameter":"Inc_Per_Txn","Value":1000},
        {"Parameter":"Min_Prod","Value":3},
        {"Parameter":"Sent_Table_Max","Value":4},
    ])
    nursery_grid  = pd.DataFrame([
        {"Min_CMR_Pct":80,"Mult":1.0},{"Min_CMR_Pct":60,"Mult":0.6},
        {"Min_CMR_Pct":50,"Mult":0.35},{"Min_CMR_Pct":0,"Mult":0.0},
    ])
    nursery_table = pd.DataFrame([
        {"Sent":4,"Min_Recd":2,"Mult":1.0},{"Sent":3,"Min_Recd":2,"Mult":1.0},
        {"Sent":2,"Min_Recd":1,"Mult":1.0},{"Sent":1,"Min_Recd":0,"Mult":1.0},
        {"Sent":0,"Min_Recd":0,"Mult":0.0},
    ])
    bm_csd = pd.DataFrame([{"Min_Ach_Pct":100,"Inc":20000},{"Min_Ach_Pct":95,"Inc":15000},{"Min_Ach_Pct":85,"Inc":10000}])
    bm_kcd = pd.DataFrame([{"Min_Ach_Pct":100,"Inc":20000},{"Min_Ach_Pct":90,"Inc":15000},{"Min_Ach_Pct":80,"Inc":10000}])
    ch_csd = pd.DataFrame([{"Min_Ach_Pct":100,"Inc":25000},{"Min_Ach_Pct":90,"Inc":20000}])
    ch_kcd = pd.DataFrame([{"Min_Ach_Pct":100,"Inc":25000},{"Min_Ach_Pct":90,"Inc":20000}])
    bt_bm_csd = pd.DataFrame([{"Deal_Size":"3L+","Min_Lakh":300,"Per_Deal":15000},{"Deal_Size":"2L+","Min_Lakh":200,"Per_Deal":10000},{"Deal_Size":"1L+","Min_Lakh":100,"Per_Deal":5000}])
    bt_ch_csd = pd.DataFrame([{"Deal_Size":"3L+","Min_Lakh":300,"Per_Deal":25000},{"Deal_Size":"2L+","Min_Lakh":200,"Per_Deal":15000},{"Deal_Size":"1L+","Min_Lakh":100,"Per_Deal":7500}])
    bt_bm_kcd = pd.DataFrame([{"Deal_Size":"10L+","Min_Lakh":1000,"Per_Deal":30000},{"Deal_Size":"8L+","Min_Lakh":800,"Per_Deal":20000},{"Deal_Size":"5L+","Min_Lakh":500,"Per_Deal":10000},{"Deal_Size":"3L+","Min_Lakh":300,"Per_Deal":5000}])
    bt_ch_kcd = pd.DataFrame([{"Deal_Size":"10L+","Min_Lakh":1000,"Per_Deal":50000},{"Deal_Size":"8L+","Min_Lakh":800,"Per_Deal":30000},{"Deal_Size":"5L+","Min_Lakh":500,"Per_Deal":15000},{"Deal_Size":"3L+","Min_Lakh":300,"Per_Deal":7500}])
    bm_csd_target = pd.DataFrame([{"Parameter":"Collection_Target","Value":0}])

    return {
        "CSD_Milestones":csd_milestones, "CSD_Targets":csd_targets,
        "CSD_CMR_Mult":csd_cmr_mult, "CSD_Params":csd_params,
        "CSD_Spot_2_6":csd_spot_2_6, "CSD_Spot_7_12":csd_spot_7_12, "CSD_Spot_20_30":csd_spot_20_30,
        "KCD_Milestones":kcd_milestones, "KCD_Targets":kcd_targets,
        "KCD_Spot_1_12":kcd_spot_1_12, "KCD_Spot_20_30":kcd_spot_20_30,
        "KCD_25Cr_Spot_1_12":kcd_25cr_1_12, "KCD_25Cr_Spot_20_30":kcd_25cr_20_30,
        "KCD_RM_Spot_1_12":kcd_rm_1_12, "KCD_RM_Spot_20_30":kcd_rm_20_30,
        "KCD_SS_Mult":kcd_ss_mult, "KCD_Params":kcd_params,
        "Nursery_Params":nursery_params, "Nursery_CMR_Grid":nursery_grid,
        "Nursery_CMR_Table":nursery_table,
        "BM_CSD_Slabs":bm_csd, "BM_KCD_Slabs":bm_kcd,
        "CH_CSD_Slabs":ch_csd, "CH_KCD_Slabs":ch_kcd,
        "BT_BM_CSD":bt_bm_csd, "BT_CH_CSD":bt_ch_csd,
        "BT_BM_KCD":bt_bm_kcd, "BT_CH_KCD":bt_ch_kcd,
        "BM_CSD_Target":bm_csd_target,
    }


def load_ta_slab_config(f):
    defs = build_ta_slab_config()
    if f is None: return defs
    # Read bytes once so both ExcelFile and read_excel can use it
    _b = f.getvalue() if hasattr(f, "getvalue") else f.read()
    xl = pd.ExcelFile(_io.BytesIO(_b))
    cfg = {}
    for k, dd in defs.items():
        if k in xl.sheet_names:
            df = pd.read_excel(_io.BytesIO(_b), sheet_name=k, header=1).dropna(how="all")
            cfg[k] = df
        else:
            cfg[k] = dd
    return cfg


def parse_slabs(cfg):
    def rows(k):   return cfg[k].to_dict("records")
    def param(k,p,d):
        df = cfg.get(k, pd.DataFrame())
        if len(df)==0: return d
        r = df[df.iloc[:,0].astype(str)==p]
        return float(r.iloc[0,1]) if len(r)>0 else d

    return {
        "csd_milestones":sorted(rows("CSD_Milestones"),key=lambda r:-r.get("Min_Ach_Pct",0)),
        "csd_targets":   {r["Vintage"]:float(r["Target_PCDV"]) for r in rows("CSD_Targets")},
        "csd_cmr_mult":  sorted(rows("CSD_CMR_Mult"),key=lambda r:-r.get("Min_CMR_Ach_Pct",0)),
        "csd_incr_rate": param("CSD_Params","Incr_Rate_Pct",3.0)/100,
        "csd_cmr_tgt":   param("CSD_Params","CMR_Target_Pct",40.0),
        "csd_spot_2_6":  sorted(rows("CSD_Spot_2_6"),  key=lambda r:-r.get("Txn",0)),
        "csd_spot_7_12": sorted(rows("CSD_Spot_7_12"), key=lambda r:-r.get("Txn",0)),
        "csd_spot_20_30":sorted(rows("CSD_Spot_20_30"),key=lambda r:-r.get("Txn",0)),
        "kcd_milestones":sorted(rows("KCD_Milestones"),key=lambda r:-r.get("Min_Ach_Pct",0)),
        "kcd_target_map": {(r.get("Group","KCD"),r["Vintage"]):({"KCD":{"0-270":7100,"270-2Yr":8100,"2Yr+":9100},
                        "KCD-25cr":{"0-270":6480,"270-2Yr":9100,"2Yr+":10000}}.get(
                        r.get("Group","KCD"),{}).get(r["Vintage"], float(r["Target_PCDV"]))) for r in rows("KCD_Targets")},
        "kcd_spot_1_12": sorted(rows("KCD_Spot_1_12"), key=lambda r:-r.get("PCDV",0)),
        "kcd_spot_20_30":sorted(rows("KCD_Spot_20_30"),key=lambda r:-r.get("PCDV",0)),
        "kcd_25cr_1_12": sorted(rows("KCD_25Cr_Spot_1_12"),key=lambda r:-r.get("PCDV",0)),
        "kcd_25cr_20_30":sorted(rows("KCD_25Cr_Spot_20_30"),key=lambda r:-r.get("PCDV",0)),
        "kcd_rm_1_12":   sorted(rows("KCD_RM_Spot_1_12"), key=lambda r:-r.get("PCDV",0)),
        "kcd_rm_20_30":  sorted(rows("KCD_RM_Spot_20_30"),key=lambda r:-r.get("PCDV",0)),
        "kcd_ss_mult":   sorted(rows("KCD_SS_Mult"),key=lambda r:-r.get("Min_SS_CMR_Pct",0)),
        "kcd_min_ss_sent":int(param("KCD_Params","Min_SS_Sent",3)),
        "nursery_inc_per_txn":int(param("Nursery_Params","Inc_Per_Txn",1000)),
        "nursery_min_prod":   int(param("Nursery_Params","Min_Prod",3)),
        "nursery_sent_max":   int(param("Nursery_Params","Sent_Table_Max",4)),
        "nursery_grid": sorted(rows("Nursery_CMR_Grid"), key=lambda r:-r.get("Min_CMR_Pct",0)),
        "nursery_table":{r["Sent"]:(r["Min_Recd"],r["Mult"]) for r in rows("Nursery_CMR_Table")},
        "bm_csd": sorted(rows("BM_CSD_Slabs"),key=lambda r:-r.get("Min_Ach_Pct",0)),
        "bm_kcd": sorted(rows("BM_KCD_Slabs"),key=lambda r:-r.get("Min_Ach_Pct",0)),
        "ch_csd": sorted(rows("CH_CSD_Slabs"),key=lambda r:-r.get("Min_Ach_Pct",0)),
        "ch_kcd": sorted(rows("CH_KCD_Slabs"),key=lambda r:-r.get("Min_Ach_Pct",0)),
        "bt_bm_csd":sorted(rows("BT_BM_CSD"),key=lambda r:-r.get("Min_Lakh",0)),
        "bt_ch_csd":sorted(rows("BT_CH_CSD"),key=lambda r:-r.get("Min_Lakh",0)),
        "bt_bm_kcd":sorted(rows("BT_BM_KCD"),key=lambda r:-r.get("Min_Lakh",0)),
        "bt_ch_kcd":sorted(rows("BT_CH_KCD"),key=lambda r:-r.get("Min_Lakh",0)),
    }


@st.cache_data(show_spinner=False)
def make_may_slab_excel():
    """Return May 2026 slab config — same structure as April but labelled May.
    Slab numbers are April defaults; upload the Excel to update with May actuals."""
    import shutil, openpyxl, io as _bio
    from openpyxl.styles import Font, PatternFill
    # Build from April defaults, override labels to May
    apr_bytes = make_slab_excel()
    buf = _bio.BytesIO(apr_bytes)
    wb = openpyxl.load_workbook(buf)
    # Update any Apr references in cell A1 of each sheet
    ORANGE = PatternFill("solid", fgColor="FFA500")
    for sh in wb.sheetnames:
        ws = wb[sh]
        c = ws.cell(1,1)
        if c.value and "Apr" in str(c.value):
            c.value = str(c.value).replace("Apr","May")
            c.fill = ORANGE
    # Add/update the notes sheet
    note_sh = wb["MAY_SPOT_NOTES"] if "MAY_SPOT_NOTES" in wb.sheetnames else wb.create_sheet("MAY_SPOT_NOTES")
    note_sh.cell(1,1).value = "May 2026 — Slab numbers are April defaults. Update each sheet with May actuals once available."
    note_sh.cell(1,1).font = Font(bold=True)
    out = _bio.BytesIO(); wb.save(out); return out.getvalue()

@st.cache_data(show_spinner=False)
def make_slab_excel():
    cfg = build_ta_slab_config()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
        hf = w.book.add_format({"bold":True,"bg_color":"#1F4E79","font_color":"#FFFFFF","border":1})
        nf = w.book.add_format({"italic":True,"font_color":"#595959"})
        for sh, df in cfg.items():
            df.to_excel(w, sheet_name=sh, index=False, startrow=1)
            ws = w.sheets[sh]
            ws.set_column(0, len(df.columns)-1, 22)
            for ci,col in enumerate(df.columns): ws.write(1,ci,col,hf)
            ws.write(0,0,f"TA Slab Config | {sh} | Edit values below, do NOT rename columns.",nf)


    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# RECEIPT ENRICHMENT  (generates sir's extra columns from raw receipt file)
# ──────────────────────────────────────────────────────────────────────────────
_SS_PLUS_PRODS = {"IM Star","IM Star Pro","IM Leader","IM Leader Pro",
                  "Preferred Star","Preferred Leader","IL","Preferred IL",
                  "Adv IM SS Pro","Preferred Star Addon","Preferred Leader Addon"}
_IM_STAR_LEADER_PRODS = {"IM Star","IM Star Pro","IM Leader","IM Leader Pro"}
INSTA_PRODUCTS = {"IM InstaDiamond","IM InstaGold","IM InstaPlatinum",
                  "IM insta Diamond","IM Insta Renewal",
                  "Lead Manager Pro Gold","Lead Manager Pro Platinum"}
_IM_STAR_PRO_PRODS    = {"IM Star Pro","IM Leader Pro","Adv IM SS Pro"}

# Hierarchy lookup built from structure (populated in load_ta_structure call)
_HIER = {}   # {exec_eid: {L2 ID, L2 Name, L3 ID, L3 Name, L4 ID, L4 Name, L5 ID, L5 Name, L6 ID, L6 Name}}

@st.cache_data(show_spinner=False)
def enrich_receipt_data(rec_df, structure_result=None):
    """
    Add sir's extra columns to the raw receipt file.
    All column logic reverse-engineered from sir's TA Base+Spot Incentive Scheme Apr26 file.
    """
    df = rec_df.copy()

    # ── Date columns: confirmed bins from sir ─────────────────────────────────
    # Week: WK-1=1-9, WK-2=10-16, WK-3=17-23, WK-4=24+
    # FNT:  FNT-1=1-16, FNT-2=17+
    dtc = find_col(df, ["Entry Date","Receipt Date","Clear Date"])
    if dtc:
        dates = pd.to_datetime(df[dtc], errors="coerce")
        day   = dates.dt.day.fillna(0).astype(int)
        df["Day"]  = day
        df["Week"] = day.apply(lambda d:
            "WK-1" if d<=9 else "WK-2" if d<=16 else "WK-3" if d<=23 else "WK-4")
        df["FNT"]  = day.apply(lambda d: "FNT-1" if d<=16 else "FNT-2")
    df["WK-1"]=None; df["WK-2"]=None; df["WK-3"]=None
    df["WK-4"]=None; df["WK-5"]=None; df["FNT-1"]=None; df["FNT-2"]=None

    # ── Rem (tag) column ──────────────────────────────────────────────────────
    rem_c = find_col(df, ["Rem","Rnl Remarks","Deal Remarks"])
    rem   = df[rem_c].astype(str).str.strip() if rem_c else pd.Series("", index=df.index)

    # ── Total Sale: 1 only for Upsell-NR and Upsell-Ren (confirmed from sir) ──
    df["Total Sale"] = (rem.isin(["Upsell-NR","Upsell-Ren"])).astype(int)

    # ── Prod.1: productivity weight — 1.0 for non-WIP/Balance/TDS rows ────────
    _excl = {"WIP","Balance","TDS","NACH/ECS","NR","Addon"}
    df["Prod.1"] = (~rem.isin(_excl)).astype(float)

    # ── CMR-C+1-C+2: Renewal rows get CMR tag (sequence-based) ───────────────
    # Renewal rows: CMR/CMR+1/CMR+2/... based on renewal sequence
    # Non-renewal rows: "Others"  (we set all to Others; CMR sequence requires renewal join)
    df["CMR-C+1-C+2"] = "Others"
    # Renewal rows that are tagged as Renewal → CMR (approximation; sir uses renewal month seq)
    _is_rnl = rem == "Renewal"
    df.loc[_is_rnl, "CMR-C+1-C+2"] = "CMR"

    # ── AMR: Yes only when Rem=="Renewal" AND it's an auto-renewal ────────────
    # Sir: AMR=Yes for ~787 Renewal rows (auto-renewal flag from renewal file)
    # We default to No; could be enhanced with renewal file join
    df["AMR"] = "No"

    # ── Sale Mapping: Yes when Rem is any of Upsell-NR, Upsell-Ren, Renewal, NR
    df["Sale Mapping"] = rem.apply(lambda r: "Yes" if r in {"Upsell-NR","Upsell-Ren","Renewal","NR"} else None)

    # ── Renewal Map: "Renewal" when Rem is Renewal or Upsell-Ren ─────────────
    df["Renewal Map"]  = rem.apply(lambda r: "Renewal" if r in {"Renewal","Upsell-Ren","Upsell-NR"} else None)

    # ── Collection: Yes if Status==Cleared ───────────────────────────────────
    status_c = find_col(df, ["Status","Receipt Status"])
    if status_c:
        cleared = df[status_c].astype(str).str.upper().str.strip() == "CLEARED"
        df["Collection"] = cleared.apply(lambda x: "Yes" if x else "No")
    else:
        df["Collection"] = "Yes"
    df["Final Status"] = "Tagged"

    # ── IM Star/Leader ────────────────────────────────────────────────────────
    prod_c  = find_col(df, ["Prod","Product","Tagged Services Name"])
    prod    = df[prod_c].astype(str).str.strip() if prod_c else pd.Series("", index=df.index)
    df["IM Star/Leader"] = prod.apply(
        lambda p: "Yes" if any(k in p for k in _IM_STAR_LEADER_PRODS) else "No")
    df["IM Star Pro+ New Sale"] = prod.apply(
        lambda p: "Yes" if any(k in p for k in _IM_STAR_PRO_PRODS) else "No")
    # Insta flag: 0.5 productivity weight
    _insta_mask = prod.apply(lambda p: any(k.upper() in p.upper() for k in ["INSTA","Lead Manager Pro"]))
    # Override Prod.1: Insta products get 0.5 weight
    if "Prod.1" in df.columns:
        df.loc[_insta_mask & (df["Prod.1"]>0), "Prod.1"] = 0.5

    # ── KCD-New Sale: Yes only for non-Balance/TDS/WIP/Renewal upsell rows ────
    # Sir: KCD-New Sale='Yes' for specific new sale products only
    # 0 for Balance/WIP/Renewal/NR; 'Yes' only for confirmed new KCD upsell products
    _kcd_new_rems = {"Upsell-NR","Upsell-Ren"}
    df["KCD-New Sale"] = rem.apply(lambda r: "Yes" if r in _kcd_new_rems else 0)

    # ── MDC-WS -Anurag: based on CustType column (NOT product name) ──────────
    # CustType mapping (confirmed from sir):
    # STAR, FREELIST, FCP, CATALOG, BL Paid* → MDC
    # LEADER → Leader   |  ExportTS, TSCATALOG → IVE/MDC-TS  |  TSCATALOG → MDC-TS
    cust_c = find_col(df, ["CustType","Cust Type","CUST TYPE"])
    if cust_c:
        cust = df[cust_c].astype(str).str.strip().str.upper()
        def _mdc_cat(ct):
            if "LEADER" in ct:                           return "Leader"
            if ct in {"EXPORTTS","EXPORT TS"}:           return "IVE"
            if ct == "TSCATALOG":                        return "MDC-TS"
            if any(k in ct for k in
                   {"STAR","FCP","CATALOG","FREELIST","BL PAID","MFCP","VGFCP","QGFCP"}):
                return "MDC"
            return "Other"
        df["MDC-WS -Anurag"] = cust.apply(_mdc_cat)
    else:
        # Fallback to product-based approximation
        df["MDC-WS -Anurag"] = prod.apply(lambda p:
            "Leader" if "Leader" in p else "IVE" if "IVE" in p or "Exporter" in p
            else "MDC-TS" if "Catalog" in p else "MDC")

    # ── Deal Val columns ──────────────────────────────────────────────────────
    wt_c  = find_col(df, ["WT AMT","WT_AMT","WT Amt(A)"])
    ra_c  = find_col(df, ["Receipt Amount"])
    dv_c  = find_col(df, ["Deal Value","Deal Val","Deal Val (WT)"])
    wot_c = find_col(df, ["Deal Val (WOT)"])

    wt_arr  = pd.to_numeric(df[wt_c],  errors="coerce").fillna(0) if wt_c  else pd.Series(0.0, index=df.index)
    dv_arr  = pd.to_numeric(df[dv_c],  errors="coerce").fillna(0) if dv_c  else pd.Series(0.0, index=df.index)
    wot_arr = pd.to_numeric(df[wot_c], errors="coerce").fillna(0) if wot_c else wt_arr
    ra_arr  = pd.to_numeric(df[ra_c],  errors="coerce").fillna(0) if ra_c  else wt_arr

    df["Deal Val"]     = dv_arr if dv_c else ra_arr
    df["Deal Val (WOT)"] = wot_arr
    # DV: use Deal Val (WOT) when non-zero, else Receipt Amount (not WT AMT)
    # For TDS/Balance rows: DV comes from the companion sale via GLUSER grouping
    df["DV"] = wot_arr.where(wot_arr > 0, ra_arr)

    # ── Big Ticket Slab: based on DV ─────────────────────────────────────────
    dv_val = df["DV"]
    def _bt(v):
        if v >= 500000: return "5L+"
        if v >= 300000: return "3L+"
        if v >= 200000: return "2L+"
        if v >= 100000: return "1L+"
        return 0
    df["Big Ticket-Slab"] = dv_val.apply(_bt)

    # ── Misc defaults ─────────────────────────────────────────────────────────
    df["Tue/False"]     = False
    df["Base to Listing"] = "No"
    df["MYR/F"]         = ""
    df["Correction"]    = ""; df["Reason"] = ""
    df["NA"]=""; df["NA.1"]=""; df["NA.2"]=""
    # Ensure Big Ticket-Slab is string (not mixed int/str)
    df["Big Ticket-Slab"] = df["Big Ticket-Slab"].astype(str).replace({"0":"0","nan":"0"})
    df["KCD-New Sale"]    = df["KCD-New Sale"].astype(str)
    df["Total Sale"]      = pd.to_numeric(df["Total Sale"], errors="coerce").fillna(0).astype(int)
    df["Prod.1"]          = pd.to_numeric(df["Prod.1"],    errors="coerce").fillna(0).astype(float)

    # ── Hierarchy from structure ──────────────────────────────────────────────
    ec = find_col(df, ["Sales Exec ID","EMP ID"])
    # Build EID lookup (int → str key) for hierarchy columns
    _hier_map = {}
    if ec and structure_result:
        for _eid_raw in df[ec].dropna().unique():
            _k = str(_eid_raw).split(".")[0].strip()
            if _k in structure_result:
                _hier_map[_k] = structure_result[_k]
    for h_key in ["L2 ID","L2 Name","L3 ID","L3 Name","L4 ID","L4 Name","L5 ID","L5 Name","L6  ID","L6 Name"]:
        if _hier_map:
            df[h_key] = df[ec].astype(str).str.split(".").str[0].str.strip().map(
                lambda e: _hier_map.get(e, {}).get(h_key, ""))
        else:
            df[h_key] = ""

    return df


@st.cache_data(show_spinner=False)
def enrich_refund_data(ref_df, structure_result=None):
    """Add sir's extra columns to the raw refund file."""
    df = ref_df.copy()
    dtc = find_col(df, ["Clear Date","Date","Month"])
    if dtc:
        dates = pd.to_datetime(df[dtc], errors="coerce")
        day   = dates.dt.day.fillna(0).astype(int)
        df["Day.1"] = day
        df["Week"]  = day.apply(lambda d: "WK-1" if d<=7 else "WK-2" if d<=14 else "WK-3" if d<=21 else "WK-4" if d<=28 else "WK-5")
        df["FNT"]   = day.apply(lambda d: "FNT-1" if d<=15 else "FNT-2")
    df["WK-1"]=None;df["WK-2"]=None;df["WK-3"]=None;df["WK-4"]=None;df["WK-5"]=None
    df["FNT-1"]=None;df["FNT-2"]=None
    df["Act amt"]    = pd.to_numeric(df.get("WT Amount", df.get("WT AMT",0)), errors="coerce").fillna(0)
    df["Correction"] = None; df["Reason.1"] = None
    df["Date"]       = df.get(dtc, None)
    # Hierarchy
    ec = find_col(df, ["Sales Ex. ID","Sales Exec ID","EMP ID"])
    if ec and structure_result:
        for h_key in ["L2 ID","L2 Name","L3 ID","L3 Name","L4 ID","L4 Name","L5 ID","L5 Name","L6  ID","L6 Name"]:
            df[h_key] = df[ec].astype(str).str.split(".").str[0].str.strip().map(
                lambda e, hk=h_key: structure_result.get(e, {}).get(hk, ""))
    else:
        for h_key in ["L2 ID","L2 Name","L3 ID","L3 Name","L4 ID","L4 Name","L5 ID","L5 Name","L6  ID","L6 Name"]:
            df[h_key] = ""
    return df


@st.cache_data(show_spinner=False)
def enrich_renewal_data(rnl_df, structure_result=None):
    """Add sir's extra columns to the raw renewal file."""
    df = rnl_df.copy()
    # Date of receipt
    dtc = find_col(df, ["Received Date","Date","Clear Date"])
    df["Received Date"] = df.get(dtc, None)
    # SS+ flag
    prod_c = find_col(df, ["WS/MDC Main","Prod","Product"])
    if prod_c:
        prod = df[prod_c].astype(str).str.strip()
        df["SS+"] = prod.apply(lambda p: "Yes" if any(k in p for k in _SS_PLUS_PRODS) else "No")
    else:
        df["SS+"] = "No"
    df["MDC-1 Sent"] = "No"
    df["Pending AMR"] = "No"
    df["Month.1"]    = df.get("Month", None)
    df["Reten. Anurag"] = None
    # Hierarchy
    ec = find_col(df, ["EMP ID","Sales Exec ID","Employee ID","L1"])
    if ec and structure_result:
        for h_key in ["L2 ID","L2 Name","L3 ID","L3 Name","L4 ID","L4 Name","L5 ID","L5 Name","L6  ID","L6 Name"]:
            df[h_key] = df[ec].astype(str).str.split(".").str[0].str.strip().map(
                lambda e, hk=h_key: structure_result.get(e, {}).get(hk, ""))
    else:
        for h_key in ["L2 ID","L2 Name","L3 ID","L3 Name","L4 ID","L4 Name","L5 ID","L5 Name","L6  ID","L6 Name"]:
            df[h_key] = ""
    return df


# ──────────────────────────────────────────────────────────
# UTILITY
# ──────────────────────────────────────────────────────────
_FIND_COL_CACHE = {}
def find_col(df, candidates):
    """Cached column finder — memoises per (col_tuple, candidates_tuple)."""
    key = (tuple(df.columns), tuple(candidates))
    if key in _FIND_COL_CACHE: return _FIND_COL_CACHE[key]
    def n(s): return " ".join(str(s).lower().split())
    nm = {n(c):c for c in df.columns}
    result = None
    for c in candidates:
        if n(c) in nm: result = nm[n(c)]; break
    _FIND_COL_CACHE[key] = result
    return result

@st.cache_data(show_spinner=False)
def load_excel(bts:bytes, fname:str)->pd.DataFrame:
    ext = fname.lower().rsplit(".",1)[-1] if "." in fname else "xlsx"
    buf = io.BytesIO(bts)
    try:
        return pd.read_excel(buf, engine="pyxlsb" if ext=="xlsb" else "openpyxl")
    except Exception:
        return pd.DataFrame()

def _rf(f):
    if f is None: return pd.DataFrame()
    return load_excel(f.getvalue(), f.name)

def _sf(v, d=0.0):
    try:
        fv=float(v); return fv if fv==fv else d
    except: return d

def _to_date(v):
    import datetime as dt
    if v is None: return None
    if isinstance(v,dt.date) and not isinstance(v,dt.datetime): return v
    if isinstance(v,dt.datetime): return v.date()
    if isinstance(v,pd.Timestamp): return v.date()
    try:
        fv=float(v)
        if 30000<fv<60000: return EXCEL_EPOCH+dt.timedelta(days=int(fv))
    except: pass
    try: return pd.to_datetime(str(v)).date()
    except: return None


# ──────────────────────────────────────────────────────────
# STRUCTURE LOADER
# ──────────────────────────────────────────────────────────
def load_ta_structure(f):
    if f is None: return {}
    try:
        # Normalise to bytes so we can reuse without seek issues
        if isinstance(f, (str, bytes.__class__)) and not isinstance(f, bytes):
            _raw = open(f, "rb").read()   # file path string
        elif isinstance(f, bytes):
            _raw = f
        else:
            _raw = f.read() if hasattr(f, "read") else f.getvalue()
        xl = pd.ExcelFile(_io.BytesIO(_raw))
        norms = [s.strip().upper() for s in xl.sheet_names]
        sh = next((xl.sheet_names[i] for i,s in enumerate(norms)
                   if s in ("FSF_TA","TA","STRUCT","STRUCTURE"," FSF_TA")), None)
        if sh is None:
            return {}
        df = pd.read_excel(_io.BytesIO(_raw), sheet_name=sh)
        df.columns = [str(c).strip() for c in df.columns]
    except Exception:
        return {}

    def gc(candidates): return find_col(df, candidates)
    ec  = gc(["Employee ID","Emp ID","EmpID"])
    nc  = gc(["Employee Name"])
    dc  = gc(["Designation"])
    vc  = gc(["IIL Vertical Name","Vertical","IIL Vertical"])
    jc  = gc(["Joining Date","DOJ"])  # Move/Join Date is NaT for TA employees
    fgc = gc(["Final Group","FinalGroup","Vintage"])
    vbc = gc(["Vintage Bucket"])
    cac = gc(["Client-A","Client A"])
    ccc = gc(["Client-C","Client C"])
    agc = gc(["Aeging","Ageing","Aging"])
    hcc = gc(["HC"])
    grc = gc(["Group"])
    rmc = gc(["Remarks"])
    l2i = gc(["L2 ID","L2ID"]); l2n = gc(["L2 Name","L2Name"])
    l3i = gc(["L3 ID","L3ID"]); l3n = gc(["L3 Name","L3Name"])
    l4i = gc(["L4 ID","L4ID"]); l4n = gc(["L4 Name","L4Name"])
    l5i = gc(["L5 ID","L5ID"]); l5n = gc(["L5 Name","L5Name"])
    l6i = gc(["L6 ID","L6  ID","L6ID"]); l6n = gc(["L6 Name","L6Name"])
    # Product portfolio columns (from structure for Nursery sheet)
    mdc_c  = gc(["MDC"]); star_c = gc(["Star"]); ldr_c  = gc(["Leader"])
    wsm_c  = gc(["WS-M"]); wsa_c  = gc(["WS-A"]); ive_c  = gc(["IVE"])
    # Renewal product columns (MDC.1 etc in structure)
    mdc1_c = gc(["MDC.1"]); star1_c= gc(["Star.1"]); ldr1_c = gc(["Leader.1"])
    wsm1_c = gc(["WS-M.1"]); wsa1_c = gc(["WS-A.1"]); ive1_c = gc(["IVE.1"])
    email_c= gc(["Email Id"]); loc_c  = gc(["Location","New Location/ROI Location"])
    dept_c = gc(["Department"])
    cstat_c= gc(["Client Status","ClientStatus"])

    result = {}
    for _, row in df.iterrows():
        if not ec: break
        eid = str(row[ec]).strip().split(".")[0]
        if not eid or eid.lower() in ("nan","none",""): continue

        vraw = str(row[vc]).strip().upper() if vc else ""
        # Only include Tele Annual employees - must have TELE in vertical name
        if "TELE" not in vraw: continue
        if "CSD" in vraw:   vert = "CSD"
        elif "KCD" in vraw: vert = "KCD"
        else: continue

        desig  = str(row[dc]).strip().upper() if dc else "L1"
        vint   = str(row[fgc]).strip() if fgc else "91-270D"
        vint = "90-270"  # placeholder; overridden by ageing in calc loop

        ageing = 0; fy_ageing = 0
        # Compute both CALC_DATE ageing and FY_START (Apr 1) ageing
        # Sir's Aeging column = FY ageing (days from Apr 1)
        if agc and str(row[agc]).strip() not in ("nan","None",""):
            try: fy_ageing = int(float(row[agc]))
            except: pass
        jd_val = row[jc] if jc else None
        jd = _to_date(jd_val) if jd_val is not None else None
        if jd and str(jd) not in ("NaT","nan","None",""):
            try:
                import datetime as _dt
                jd_date = jd.date() if isinstance(jd, _dt.datetime) else jd
                ageing    = max(0, (CALC_DATE - jd_date).days)  # from Apr 30
                if fy_ageing == 0:
                    fy_ageing = max(0, (FY_START  - jd_date).days)  # from Apr 1
            except Exception:
                ageing = fy_ageing = 0

        def sv(c): return str(row[c]).strip() if c else ""
        def nv(c): return _sf(row[c]) if c else 0.0
        def id_sv(c): return str(row[c]).strip().split(".")[0] if c else ""

        _nurs_bkt = "60D" if str(eid).strip() in _60D_NURSERY_EIDS else "60-90"

        result[eid] = {
            "Name":     sv(nc),   "Designation": desig,
            "Vertical": vert,     "Vintage":     vint,
            "Ageing":   ageing,   "FY_Ageing": fy_ageing,  "Joining Date":_to_date(row[jc]) if jc else None,
            "Client-A": nv(cac),  "Client-C":    nv(ccc),
            "HC":       nv(hcc),  "Group":       sv(grc),
            "Remarks":  sv(rmc),  "Email Id":    sv(email_c),
            "_nurs_bucket": _nurs_bkt,
            "Department": sv(dept_c),
            "Client_Status_Eligible": (str(row[cstat_c]).strip().upper()=="YES") if cstat_c else True,
            "Location": sv(loc_c),
            "MDC":int(nv(mdc_c)),"Star":int(nv(star_c)),"Leader":int(nv(ldr_c)),
            "WS-M":int(nv(wsm_c)),"WS-A":int(nv(wsa_c)),"IVE":int(nv(ive_c)),
            "MDC.1":int(nv(mdc1_c)),"Star.1":int(nv(star1_c)),"Leader.1":int(nv(ldr1_c)),
            "WS-M.1":int(nv(wsm1_c)),"WS-A.1":int(nv(wsa1_c)),"IVE.1":int(nv(ive1_c)),
            "L2 ID":id_sv(l2i),"L2 Name":sv(l2n),
            "L3 ID":id_sv(l3i),"L3 Name":sv(l3n),
            "L4 ID":id_sv(l4i),"L4 Name":sv(l4n),
            "L5 ID":id_sv(l5i),"L5 Name":sv(l5n),
            "L6 ID":id_sv(l6i),"L6 Name":sv(l6n),
        }

    # Aggregate Client-A and HC for all management levels from L1 data
    if ec and cac:
        try:
            df["_eid"] = df[ec].astype(str).str.split(".").str[0].str.strip()
            df[cac] = pd.to_numeric(df[cac], errors="coerce").fillna(0)
            # Only L1 rows for aggregation
            l1_df = df[df[dc].astype(str).str.strip().str.upper()=="L1"] if dc else df
            # FY ageing for nursery count (Apr 1 reference)
            import datetime as _dt2
            _FY = _dt2.date(2026, 4, 1)
            _jds = pd.to_datetime(df[jc], errors="coerce").dt.date if jc else None
            _mjd_nm = next((c for c in df.columns if "move" in c.lower() and "join" in c.lower()), None)
            _mjds = pd.to_datetime(df[_mjd_nm], errors="coerce").dt.date if _mjd_nm else None
            def _calc_fy(jd, mjd):
                try:
                    if mjd and str(mjd) not in ("NaT","nan","None",""):
                        d = (_FY - mjd).days
                        if 0 <= d <= 90: return True
                    if jd and str(jd) not in ("NaT","nan","None",""):
                        return 0 <= (_FY - jd).days <= 90
                except: pass
                return False
            _eid_col = ec  # Employee ID column
            df["_is_nurs"] = [
                (str(row_eid).strip().split(".")[0] in _60D_NURSERY_EIDS or _calc_fy(j, m))
                for row_eid, j, m in zip(
                    df[_eid_col].tolist() if _eid_col else [""]*len(df),
                    list(_jds)  if _jds  is not None else [None]*len(df),
                    list(_mjds) if _mjds is not None else [None]*len(df))
            ]
            _ccc = gc(["Client-C","Client C","ClientC"])
            # Ensure Client-C is numeric
            if _ccc and _ccc in df.columns:
                df[_ccc] = pd.to_numeric(df[_ccc], errors="coerce").fillna(0)
            if _ccc and _ccc in df.columns:
                df[_ccc] = pd.to_numeric(df[_ccc], errors="coerce").fillna(0)

            def _agg_for(id_col, src_df=None):
                if not id_col: return {}
                src = src_df if src_df is not None else l1_df
                gk = src[id_col].astype(str).str.split(".").str[0].str.strip()
                base = src.groupby(gk).agg(ca_sum=(cac,"sum"), l1_cnt=("_eid","count"))
                if _ccc and _ccc in src.columns:
                    base["cc_sum"] = src.groupby(gk)[_ccc].sum()
                if "_is_nurs" in src.columns:
                    base["nursery_cnt"] = src.groupby(gk)["_is_nurs"].sum()
                return base.to_dict("index")

            # Build subordinates_map: {l2_eid: [l1_eid, ...]}
            subordinates_map = {}
            if l2i:
                for _, _sr in l1_df.iterrows():
                    _l2 = str(_sr[l2i]).strip().split(".")[0]
                    _l1 = str(_sr[ec]).strip().split(".")[0]
                    if _l2 and _l2 not in ("nan","none",""):
                        subordinates_map.setdefault(_l2, [])
                        if _l1 not in subordinates_map[_l2]:
                            subordinates_map[_l2].append(_l1)

            # L2 count per BM: count of L2 employees where L3 ID = BM ID
            l2_df = df[df[dc].astype(str).str.strip().str.upper()=="L2"] if dc else df
            agg_l2_cnt = _agg_for(l3i, l2_df) if l3i else {}
            agg_l2 = _agg_for(l2i)
            agg_l3 = _agg_for(l3i)
            agg_l4 = _agg_for(l4i)
            agg_l5 = _agg_for(l5i)
            for eid, emp in result.items():
                d = emp["Designation"]
                a = None
                if d=="L2":   a = agg_l2.get(eid)
                elif d=="L3":
                    a = agg_l3.get(eid)
                    l2c = agg_l2_cnt.get(eid)
                    if l2c: result[eid]["L2_Count"] = int(l2c["l1_cnt"])
                elif d=="L4": a = agg_l4.get(eid)
                elif d=="L5": a = agg_l5.get(eid)
                if a:
                    result[eid]["Client-A_Agg"] = float(a["ca_sum"])
                    result[eid]["Client-C_Agg"] = float(a.get("cc_sum", a["ca_sum"]))
                    result[eid]["HC"]            = int(a["l1_cnt"])
                    result[eid]["L1_Count"]      = int(a["l1_cnt"])
                    result[eid]["lt90_count"]    = int(a.get("nursery_cnt", 0))
                if d == "L2" and eid in subordinates_map:
                    result[eid]["l1_eids"] = subordinates_map[eid]
                if d == "L3":
                    _all=[]
                    for _l2k,_l2e in list(result.items()):
                        if _l2e.get("Designation")=="L2" and str(_l2e.get("L3 ID","")).strip()==str(eid):
                            _all.extend(subordinates_map.get(_l2k,[]))
                    if str(eid) in subordinates_map:
                        _all.extend(subordinates_map[str(eid)])
                    result[eid]["all_l1_eids"]=list(set(_all))
                
                if d == "L4":
                    # L4 CH: gather all L1s from under their L3 BMs
                    _all_l4=[]
                    for _l3k,_l3e in list(result.items()):
                        if (_l3e.get("Designation")=="L3" and
                            str(_l3e.get("L4 ID","")).strip()==str(eid)):
                            _all_l4.extend(_l3e.get("all_l1_eids",[]))
                    result[eid]["all_l1_eids"]=list(set(_all_l4))
                    # L2 count for CH
                    _l2_cnt=sum(1 for _l2k,_l2e in result.items()
                                if _l2e.get("Designation")=="L2" and
                                   str(result.get(str(_l2e.get("L3 ID","")),{}).get("L4 ID","")).strip()==str(eid))
                    result[eid]["L2_Count"]=_l2_cnt
        except Exception:
            import traceback; traceback.print_exc()
    return result


# ──────────────────────────────────────────────────────────
# MONTH FILTERING
# ──────────────────────────────────────────────────────────
def get_months(rec_df, rnl_df):
    months = set()
    dc = find_col(rec_df, ["Entry Date","Clear Date","Receipt Date"])
    if dc:
        for d in pd.to_datetime(rec_df[dc], errors="coerce").dropna():
            months.add(d.strftime("%b-%y"))
    if rnl_df is not None:
        mc = find_col(rnl_df, ["Month","MONTH"])
        if mc:
            for m in rnl_df[mc].dropna().unique():
                try:
                    p = pd.to_datetime(str(m), format="%b'%y", errors="coerce")
                    if pd.notna(p): months.add(p.strftime("%b-%y"))
                except: pass
    return sorted(months, key=lambda x: pd.to_datetime(x, format="%b-%y"))


def get_prev_month_str(sel, available_months):
    """Return previous month string for CMR+1, or None if unavailable."""
    try:
        import pandas as _pd2
        nxt = (_pd2.to_datetime(sel, format="%b-%y") + _pd2.DateOffset(months=1)).strftime("%b-%y")
        return nxt if nxt in available_months else None
    except: return None

def filter_month(rec, ref, rnl, sel):
    # sel can be "May'26", "May-26", "Apr'26" etc — normalise before parsing
    _sel_norm = str(sel).strip().replace("'", "-").replace(" ", "-")
    tgt = pd.to_datetime(_sel_norm, format="%b-%y")
    tm, ty = tgt.month, tgt.year

    def _dt(s):
        p = pd.to_datetime(s, errors="coerce")
        nums = pd.to_numeric(s, errors="coerce")
        ec = (p.dt.year==1970).sum() if p.notna().any() else 0
        if ec > p.notna().sum()*0.5:
            base = pd.Timestamp("1899-12-30")
            p = nums.apply(lambda x: base+pd.Timedelta(days=int(x)) if pd.notna(x) and x>0 else pd.NaT)
        return p

    dc = find_col(rec, ["Entry Date","Clear Date","Receipt Date"])
    r  = rec.copy()
    if dc:
        dt = _dt(r[dc]); r = r[(dt.dt.month==tm)&(dt.dt.year==ty)]

    rf = ref.copy()
    rdc = find_col(rf, ["Clear Date","Month"])
    if rdc:
        dt2 = _dt(rf[rdc]); rf = rf[(dt2.dt.month==tm)&(dt2.dt.year==ty)]

    rn = rnl.copy() if rnl is not None else None
    if rn is not None:
        mc = find_col(rn, ["Month","MONTH"])
        if mc:
            # Vectorised parse: "Apr'26" → replace ' with space → "Apr 26"
            _parsed = pd.to_datetime(
                rn[mc].astype(str).str.strip().str.replace("'", " ", regex=False),
                format="%b %y", errors="coerce"
            )
            rn = rn[(_parsed.dt.month == tm) & (_parsed.dt.year == ty)]
    return r, rf, rn


# ──────────────────────────────────────────────────────────
# CMR
# ──────────────────────────────────────────────────────────
def calc_cmr(rnl, eid_str):
    zero = {"sent":0,"recd":0,"pct":0.0,"ss_sent":0,"ss_recd":0,"ss_pct":0.0}
    if rnl is None or len(rnl)==0: return zero
    ec = find_col(rnl, ["EMP ID","Emp ID","EmpID","Employee ID"])
    sc = find_col(rnl, ["Status","STATUS"])
    pc = find_col(rnl, ["WS/MDC Main","Product","Service"])
    if not ec: return zero
    df = rnl.copy()
    df[ec] = df[ec].astype(str).str.split(".").str[0].str.strip()
    grp = df[df[ec]==eid_str]
    if len(grp)==0: return zero
    rm = grp[sc].astype(str).str.upper().str.contains("RECEIVED",na=False) if sc else pd.Series(False,index=grp.index)
    sent,recd = len(grp),int(rm.sum())
    pct = round(recd/sent*100,2) if sent>0 else 0.0
    if pc:
        ss = grp[pc].astype(str).str.upper().apply(lambda x: any(k in x for k in SS_PLUS_KW))
    else:
        ss = pd.Series(False,index=grp.index)
    ss_s = int(ss.sum()); ss_r = int((ss&rm).sum())
    return {"sent":sent,"recd":recd,"pct":pct,
            "ss_sent":ss_s,"ss_recd":ss_r,
            "ss_pct":round(ss_r/ss_s*100,2) if ss_s>0 else 0.0}


def calc_cmr_from_eids(rnl, l1_eids):
    """Aggregate CMR across all L1 EIDs for an L2 RM."""
    zero = {"sent":0,"recd":0,"pct":0.0,"ss_sent":0,"ss_recd":0,"ss_pct":0.0}
    if rnl is None or len(rnl)==0 or not l1_eids: return zero
    ec2 = find_col(rnl, ["EMP ID","Emp ID","EmpID","Employee ID"])
    sc2 = find_col(rnl, ["Status","STATUS"])
    pc2 = find_col(rnl, ["WS/MDC Main","Product","Service"])
    if not ec2: return zero
    df2 = rnl.copy()
    df2[ec2] = df2[ec2].astype(str).str.split(".").str[0].str.strip()
    grp = df2[df2[ec2].isin([str(x) for x in l1_eids])]
    if len(grp)==0: return zero
    rm2 = grp[sc2].astype(str).str.upper().str.contains("RECEIVED",na=False) if sc2 else pd.Series(False,index=grp.index)
    sent,recd = len(grp),int(rm2.sum())
    pct = round(recd/sent*100,2) if sent>0 else 0.0
    ss2 = grp[pc2].astype(str).str.upper().apply(lambda x: any(k in x for k in SS_PLUS_KW)) if pc2 else pd.Series(False,index=grp.index)
    ss_s,ss_r = int(ss2.sum()),int((ss2&rm2).sum())
    return {"sent":sent,"recd":recd,"pct":pct,"ss_sent":ss_s,"ss_recd":ss_r,
            "ss_pct":round(ss_r/ss_s*100,2) if ss_s>0 else 0.0}

def build_rnl_eid_idx(rnl):
    """Build {eid_str: sub_df} index from an already-month-filtered renewal DF."""
    idx = {}
    if rnl is None or len(rnl) == 0:
        return idx
    ec = find_col(rnl, ["EMP ID","Emp ID","EmpID","Employee ID"])
    if not ec:
        return idx
    normed = rnl[ec].astype(str).str.split(".").str[0].str.strip()
    rnl2 = rnl.copy()
    rnl2["_eid_norm"] = normed
    for eid_val, grp in rnl2.groupby("_eid_norm"):
        idx[str(eid_val)] = grp
    return idx


def _cmr_from_grp(grp, sc, pc):
    """Compute CMR stats dict from an already-filtered sub-DataFrame."""
    zero = {"sent":0,"recd":0,"pct":0.0,"ss_sent":0,"ss_recd":0,"ss_pct":0.0}
    if grp is None or len(grp) == 0:
        return zero
    rm = grp[sc].astype(str).str.upper().str.contains("RECEIVED", na=False) if sc else pd.Series(False, index=grp.index)
    sent, recd = len(grp), int(rm.sum())
    pct = round(recd / sent * 100, 2) if sent > 0 else 0.0
    ss = grp[pc].astype(str).str.upper().apply(lambda x: any(k in x for k in SS_PLUS_KW)) if pc else pd.Series(False, index=grp.index)
    ss_s = int(ss.sum()); ss_r = int((ss & rm).sum())
    return {"sent":sent,"recd":recd,"pct":pct,
            "ss_sent":ss_s,"ss_recd":ss_r,
            "ss_pct":round(ss_r/ss_s*100,2) if ss_s > 0 else 0.0}


def calc_cmr_fast_eid(eid_idx, eid_str, sc, pc):
    """O(1) CMR lookup using pre-built EID index."""
    return _cmr_from_grp(eid_idx.get(eid_str), sc, pc)


def calc_cmr_fast_eids_idx(eid_idx, l1_eids, sc, pc):
    """O(k) CMR aggregation for L2 using pre-built EID index (k = team size)."""
    frames = [eid_idx[e] for e in l1_eids if e in eid_idx]
    if not frames:
        return {"sent":0,"recd":0,"pct":0.0,"ss_sent":0,"ss_recd":0,"ss_pct":0.0}
    merged = pd.concat(frames, ignore_index=True)
    return _cmr_from_grp(merged, sc, pc)


# ── PERFORMANCE INDICES ──────────────────────────────────────
def build_renewal_index(rnl):
    if rnl is None or len(rnl)==0: return {}
    ec=find_col(rnl,["EMP ID","Emp ID","EmpID","Employee ID"])
    mc=find_col(rnl,["Month","MONTH"])
    if not ec or not mc: return {}
    rnl=rnl.copy()
    rnl["_en"]=rnl[ec].astype(str).str.split(".").str[0].str.strip()
    rnl["_mn"]=rnl[mc].astype(str).str.strip()
    idx={}
    for (e,m),g in rnl.groupby(["_en","_mn"]): idx[(str(e),str(m))]=g
    return idx

def calc_cmr_fast(rnl_idx,eid_str,month_str):
    rows_df=rnl_idx.get((eid_str,month_str),pd.DataFrame())
    zero={"sent":0,"recd":0,"pct":0.0,"ss_sent":0,"ss_recd":0,"ss_pct":0.0}
    if len(rows_df)==0: return zero
    sc=find_col(rows_df,["Status","STATUS"]); pc=find_col(rows_df,["WS/MDC Main","Product","Service"])
    rm=rows_df[sc].astype(str).str.upper().str.contains("RECEIVED",na=False) if sc else pd.Series(False,index=rows_df.index)
    sent,recd=len(rows_df),int(rm.sum())
    pct=round(recd/sent*100,2) if sent>0 else 0.0
    ss=rows_df[pc].astype(str).str.upper().apply(lambda x:any(k in x for k in SS_PLUS_KW)) if pc else pd.Series(False,index=rows_df.index)
    ss_s=int(ss.sum()); ss_r=int((ss&rm).sum())
    return {"sent":sent,"recd":recd,"pct":pct,"ss_sent":ss_s,"ss_recd":ss_r,"ss_pct":round(ss_r/ss_s*100,2) if ss_s>0 else 0.0}

def calc_cmr_fast_eids(rnl_idx,eids,month_str):
    tot={"sent":0,"recd":0,"pct":0.0,"ss_sent":0,"ss_recd":0,"ss_pct":0.0}
    for eid in eids:
        c=calc_cmr_fast(rnl_idx,eid,month_str)
        tot["sent"]+=c["sent"]; tot["recd"]+=c["recd"]; tot["ss_sent"]+=c["ss_sent"]; tot["ss_recd"]+=c["ss_recd"]
    if tot["sent"]>0:
        tot["pct"]=round(tot["recd"]/tot["sent"]*100,2)
        tot["ss_pct"]=round(tot["ss_recd"]/tot["ss_sent"]*100,2) if tot["ss_sent"]>0 else 0.0
    return tot

def calc_cmr_by_name(rnl, l2_col, name):
    zero = {"sent":0,"recd":0,"pct":0.0,"ss_sent":0,"ss_recd":0,"ss_pct":0.0}
    if rnl is None or not l2_col or not name or l2_col not in rnl.columns: return zero
    grp = rnl[rnl[l2_col].astype(str).str.strip()==name.strip()]
    if len(grp)==0: return zero
    sc = find_col(rnl, ["Status"])
    pc = find_col(rnl, ["WS/MDC Main","Product"])
    rm = grp[sc].astype(str).str.upper().str.contains("RECEIVED",na=False) if sc else pd.Series(False,index=grp.index)
    sent,recd = len(grp),int(rm.sum())
    pct = round(recd/sent*100,2) if sent>0 else 0.0
    if pc:
        ss = grp[pc].astype(str).str.upper().apply(lambda x: any(k in x for k in SS_PLUS_KW))
    else:
        ss = pd.Series(False,index=grp.index)
    ss_s=int(ss.sum()); ss_r=int((ss&rm).sum())
    return {"sent":sent,"recd":recd,"pct":pct,
            "ss_sent":ss_s,"ss_recd":ss_r,
            "ss_pct":round(ss_r/ss_s*100,2) if ss_s>0 else 0.0}


# ──────────────────────────────────────────────────────────
# RECEIPT EXTRACTION
# ──────────────────────────────────────────────────────────
def clean_receipt(df):
    """Keep only CLEARED rows with no B/C flag. Do NOT filter by Vertical -
    TA employees are identified by EMP ID matching against struct_map."""
    df = df.copy()
    bc = find_col(df,["B/C"])
    if bc: df = df[df[bc].isna()|(df[bc].astype(str).str.strip()=="")]
    sc = find_col(df,["Status","PAYMENT STATUS","Payment Status"])
    if sc:
        cl = df[df[sc].astype(str).str.upper().str.strip()=="CLEARED"]
        if len(cl)>0: df=cl
    return df


# ── RECEIPT INDEX (built once, used for all employees) ──────────────────
_REC_IDX   = {}   # {eid_str: sub-DataFrame}  keyed by Sales Exec ID
_HOD_IDX   = {}   # {eid_str: sub-DataFrame}  keyed by HOD Id
_REF_IDX   = {}   # {eid_str: sub-DataFrame}  keyed by Sales Exec ID (refund)

def build_receipt_indices(rec, ref):
    """Pre-build all receipt/refund indices. Call once after filtering by month."""
    global _REC_IDX, _HOD_IDX, _REF_IDX
    _REC_IDX = {}; _HOD_IDX = {}; _REF_IDX = {}
    ec  = find_col(rec, ["Sales Exec ID","EMP ID","Emp ID"])
    hc  = find_col(rec, ["HOD Id","Old Sales Hod ID"])
    rfc = find_col(ref, ["Sales Ex. ID","Sales Exec ID","EMP ID"])
    if ec:
        norm = rec[ec].astype(str).str.split(".").str[0].str.strip()
        for eid, grp in rec.groupby(norm): _REC_IDX[str(eid)] = grp
    if hc:
        norm_h = rec[hc].astype(str).str.split(".").str[0].str.strip()
        for eid, grp in rec.groupby(norm_h): _HOD_IDX[str(eid)] = grp
    if rfc:
        norm_r = ref[rfc].astype(str).str.split(".").str[0].str.strip()
        for eid, grp in ref.groupby(norm_r): _REF_IDX[str(eid)] = grp


def get_emp_data(rec, ref, eid_str, desig="L1", emp_name="", client_a=1, is_l2=False, l1_eids=None):
    """Fast receipt lookup using pre-built indices."""
    ec  = find_col(rec,["Sales Exec ID","EMP ID","Emp ID"])
    dvc = find_col(rec,["Deal Value","Deal Val (WT)","Deal Val"])
    wtc = find_col(rec,["Receipt Amount","WT AMT","WT_AMT"])  # Sir uses Receipt Amount
    dtc = find_col(rec,["Entry Date","Receipt Date"])  # Use Entry Date only for spot period
    upc = find_col(rec,["Unique","Upsell","UNIQUE"])
    if not ec: return {}

    # Use pre-built index for O(1) lookup
    if (desig == "L2" or is_l2) and l1_eids:
        # Aggregate L1 subordinates from index
        frames = [_REC_IDX[e] for e in l1_eids if e in _REC_IDX]
        r = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=rec.columns)
    elif desig in ("L3","L4","L5","L6"):
        r = _HOD_IDX.get(eid_str, pd.DataFrame(columns=rec.columns))
    else:
        r = _REC_IDX.get(eid_str, pd.DataFrame(columns=rec.columns))

    gross = _sf(r[wtc].fillna(0).sum()) if wtc else 0.0

    # Productivity-weighted deal value: Prod.1=1.0 regular, 0.5 Insta, 0 excluded
    # Use Prod.1 column if present (from enriched receipt), else fallback to raw DV
    _prod1_c = find_col(r, ["Prod.1","Productivity","Prod_1"]) if len(r)>0 else None
    if dvc and len(r)>0:
        _dv_arr  = pd.to_numeric(r[dvc], errors="coerce").fillna(0)
        if _prod1_c:
            _p1_arr  = pd.to_numeric(r[_prod1_c], errors="coerce").fillna(1.0)
            # Where Prod.1=0 (excluded rows), don't count DV
            _p1_arr  = _p1_arr.clip(0, 1)          # cap at 1
            dval = _sf((_dv_arr * _p1_arr).sum())  # weighted deal value
        else:
            # Fallback: exclude WIP/Balance/TDS rows using Rem column
            _rem_c2 = find_col(r, ["Rem","Rnl Remarks","Deal Remarks"])
            if _rem_c2:
                _excl_mask = r[_rem_c2].astype(str).str.strip().isin(
                    {"WIP","Balance","TDS","NACH/ECS","NR","Addon"})
                dval = _sf(_dv_arr[~_excl_mask].sum())
            else:
                dval = _sf(_dv_arr.sum())
    else:
        dval = 0.0

    # Refund — fast index lookup (aggregate team for L2/L3/L4)
    refund = 0.0
    wac = find_col(ref, ["WT Amount","WT_Amount","WT AMT"])
    if wac:
        if l1_eids:
            refund = sum(_sf(_REF_IDX[e][wac].fillna(0).sum()) for e in l1_eids if e in _REF_IDX)
        else:
            rr = _REF_IDX.get(eid_str, pd.DataFrame())
            if len(rr): refund = _sf(rr[wac].fillna(0).sum())

    net_coll = gross - refund
    client_a_eff = max(float(client_a),1)

    # Period-based deal value & spot txn counts
    fnt_dv   = {"1_12":0.0,"20_30":0.0}
    spot_txn = {"2_6":0,"7_12":0,"20_30":0}
    if dtc and len(r)>0:
        try:
            dates = pd.to_datetime(r[dtc], errors="coerce")
            nums  = pd.to_numeric(r[dtc], errors="coerce")
            if (dates.dt.year==1970).sum() > dates.notna().sum()*0.4:
                base=pd.Timestamp("1899-12-30")
                dates=nums.apply(lambda x:base+pd.Timedelta(days=int(x)) if pd.notna(x) and x>0 else pd.NaT)
            # Spot counts: productive rows only (Prod.1>0) AND new sales (Upsell-NR/Ren)
            _rem_c = find_col(r, ["Rem","Rnl Remarks","Deal Remarks","Remarks"])
            _prod1_spot = find_col(r, ["Prod.1","Productivity","Prod_1"])
            if _rem_c:
                _new_sale_mask = r[_rem_c].astype(str).str.strip().isin({"Upsell-NR","Upsell-Ren"})
            else:
                _ts_c = find_col(r, ["Total Sale","total_sale"])
                _new_sale_mask = (pd.to_numeric(r[_ts_c],errors="coerce")==1) if _ts_c else pd.Series(True, index=r.index)
            # Also restrict to productive rows (Prod.1>0) when available
            if _prod1_spot:
                _prod_mask = pd.to_numeric(r[_prod1_spot], errors="coerce").fillna(1.0) > 0
                _new_sale_mask = _new_sale_mask & _prod_mask
            days = dates.dt.day
            _spot_days = days[_new_sale_mask]  # only new sale days for spot count
            if dvc:
                dv_s = pd.to_numeric(r[dvc],errors="coerce").fillna(0)
                fnt_dv["1_12"]  = float(dv_s[days<=12].sum())
                fnt_dv["20_30"] = float(dv_s[days>=20].sum())
            # Spot periods: April=2-6/7-12/20-30, May=1-12/4th/20-30
            # p1, p2, p3 driven by CALC_DATE month
            _mo = CALC_DATE.month
            if _mo == 4:   # April
                spot_txn["2_6"]   = int(_spot_days.between(2,6).sum())
                spot_txn["7_12"]  = int(_spot_days.between(7,12).sum())
                spot_txn["20_30"] = int((_spot_days>=20).sum())
                spot_txn["1_12"] = spot_txn["2_6"] + spot_txn["7_12"]
                spot_txn["4th"]  = int((_spot_days==4).sum())
            else:          # May (and default)
                spot_txn["1_12"]  = int(_spot_days.between(1,12).sum())
                spot_txn["4th"]   = int((_spot_days==4).sum())
                spot_txn["20_30"] = int((_spot_days>=20).sum())
                spot_txn["2_6"]  = spot_txn["1_12"]
                spot_txn["7_12"] = spot_txn["4th"]
        except: pass

    # IM Star/Leader new sale counts
    im_count = 0; im_pro_count = 0; ss_spot20_30 = 0
    if len(r)>0 and dtc:
        try:
            _dts_ss = pd.to_datetime(r[dtc], errors="coerce")
            _day_ss = _dts_ss.dt.day
            # SS+ (IM Star/Leader) receipts in day 20-30 for spot multiplier
            if upc:
                _uv_ss = r[upc].fillna("").astype(str).str.upper()
                _ss_mask = _uv_ss.apply(lambda x: any(k in x for k in IM_STAR_LEADER_KW))
                ss_spot20_30 = int((_day_ss >= 20).fillna(False) & _ss_mask).sum() if hasattr((_day_ss>=20),'sum') else 0
                # safer:
                ss_spot20_30 = int(((_day_ss >= 20) & _ss_mask).sum())
        except: pass
    if upc and len(r)>0:
        uv = r[upc].fillna("").astype(str).str.upper()
        im_count = int(uv.apply(lambda x: any(k in x for k in IM_STAR_LEADER_KW)).sum())
        if dtc:
            try:
                dts = pd.to_datetime(r[dtc],errors="coerce")
                d28 = dts.dt.day >= 28
                im_pro_count = int((d28 & uv.apply(
                    lambda x: any(k in x for k in IM_STAR_PRO_KW))).sum())
            except: im_pro_count=0

    pcdv_1_12  = fnt_dv["1_12"]  / client_a_eff
    pcdv_20_30 = fnt_dv["20_30"] / client_a_eff
    pcdv_total = dval / client_a_eff

    return {
        "gross":gross,"refund":refund,"net_coll":net_coll,
        "deal_val":dval,"pcdv":pcdv_total,
        "fnt_dv":fnt_dv,"pcdv_1_12":pcdv_1_12,"pcdv_20_30":pcdv_20_30,
        "spot_txn":spot_txn,"im_count":im_count,"im_pro_count":im_pro_count,
        "ss_spot20_30":ss_spot20_30,"txns":len(r),
    }


# ──────────────────────────────────────────────────────────
# INCENTIVE CALCULATORS
# ──────────────────────────────────────────────────────────
def _milestone(pcdv, tgt, slabs):
    if tgt<=0: return 0
    ach = pcdv/tgt*100
    for r in slabs:
        if ach >= r.get("Min_Ach_Pct",r.get("Min_Achievement_Pct",0)):
            return r.get("Grid",r.get("Grid_Incentive",0))
    return 0

def _txn_spot(txn, slabs):
    for r in slabs:
        if txn >= r.get("Txn",r.get("Txn_Count",0)): return r.get("Inc",r.get("Incentive",0))
    return 0

def _pcdv_spot(pcdv, slabs):
    for r in slabs:
        if pcdv >= r.get("PCDV",r.get("PCDV_Threshold",0)): return r.get("Inc",r.get("Incentive",0))
    return 0

def _cmr_mult(cmr_pct, cmr_tgt, slabs):
    if cmr_tgt<=0: return 1.0
    ach = cmr_pct/cmr_tgt*100
    for r in slabs:
        if ach >= r.get("Min_CMR_Ach_Pct",0): return r.get("Multiplier",0)
    return 0.0

def _ss_mult(ss_pct, ss_sent, min_sent, slabs):
    if ss_sent < min_sent: return 1.0
    for r in slabs:
        if ss_pct >= r.get("Min_SS_CMR_Pct",0): return r.get("Multiplier",0.5)
    return 0.5

def _nursery_mult(sent, recd, cmr_pct, sent_max, cmr_grid, cmr_table,
                  bucket="60-90", grid_60d=None, fy_age=999, vertical="CSD"):
    """
    Official PDF Nursery Scheme — Renewal Multiplier.

    CSD Scheme (PDF 1):
      Sent=4: Recd>=2    Sent>=5 CMR grid: >=35% (0-60D), >=40% (61-90D)

    KCD/25Cr+ Scheme (PDF 2):
      Sent=4: Recd>=3    Sent>=5 CMR grid: >=60% (all periods)

    Renewal Table:
      Sent=0 → 1.0 always
      Sent=1 → 1.0 always
      Sent=2 → 1.0 if Recd>=1, else 0.0
      Sent=3 → 1.0 if Recd>=2, else 0.0
      Sent=4 → CSD: Recd>=2 | KCD: Recd>=3
      Sent>=5 → CMR% Grid (binary 1.0 or 0.0)

    All payouts are binary 1.0 or 0.0.
    """
    if vertical == "KCD":
        # KCD: Sent=4 needs Recd>=3, Sent>=5 needs CMR%>=60%
        if sent == 0: return 1.0
        if sent == 1: return 1.0
        if sent == 2: return 1.0 if recd >= 1 else 0.0
        if sent == 3: return 1.0 if recd >= 2 else 0.0
        if sent == 4: return 1.0 if recd >= 3 else 0.0   # KCD: >=3
        return 1.0 if cmr_pct >= 60.0 else 0.0           # KCD: 60% threshold (cmr_pct is %)

    # CSD (default)
    # cmr_pct is a PERCENTAGE (e.g. 8.0 = 8%) — thresholds must match
    if bucket == "60D" or fy_age <= 60:
        cmr_thresh = 35.0   # 35% for first 60D / 60D movers
    else:
        cmr_thresh = 40.0   # 40% for 61-90D

    if sent == 0: return 1.0
    if sent == 1: return 1.0
    if sent == 2: return 1.0 if recd >= 1 else 0.0
    if sent == 3: return 1.0 if recd >= 2 else 0.0
    if sent == 4: return 1.0 if recd >= 2 else 0.0   # CSD: >=2
    return 1.0 if cmr_pct >= cmr_thresh else 0.0

def _bm_milestone(ach_pct, slabs):
    for r in slabs:
        if ach_pct >= r.get("Min_Ach_Pct",0): return r.get("Inc",0)
    return 0

def _big_ticket(deal_sizes_lakh, slabs):
    """deal_sizes_lakh: list of deal sizes in lakhs. Count how many fall in each bucket."""
    total = 0
    for dv in deal_sizes_lakh:
        for r in slabs:
            if dv >= r.get("Min_Lakh",0):
                total += r.get("Per_Deal",0)
                break
    return total


# ──────────────────────────────────────────────────────────
# MAIN CALCULATOR per employee
# ──────────────────────────────────────────────────────────
def calc_employee(emp, data, cmr, S, is_25cr=False):
    """Returns dict of all incentive columns for one employee."""
    vert   = emp["Vertical"]   # "CSD" or "KCD"
    desig  = emp["Designation"]
    vint   = emp["Vintage"]
    client_a = max(float(emp.get("Client-A_Agg", emp.get("Client-A",1))),1)
    hc     = max(float(emp.get("HC",1) or 1),1)

    pcdv_total  = data.get("pcdv",0)
    pcdv_1_12   = data.get("pcdv_1_12",0)
    pcdv_20_30  = data.get("pcdv_20_30",0)
    deal_val    = data.get("deal_val",0)
    fnt_dv      = data.get("fnt_dv",{})
    spot_txn    = data.get("spot_txn",{"2_6":0,"7_12":0,"20_30":0})
    im_count    = data.get("im_count",0)
    im_pro_count= data.get("im_pro_count",0)

    cmr_pct = cmr.get("pct",0); cmr_sent = cmr.get("sent",0); cmr_recd = cmr.get("recd",0)
    ss_pct  = cmr.get("ss_pct",0); ss_sent = cmr.get("ss_sent",0); ss_recd = cmr.get("ss_recd",0)

    out = {
        "gross_coll":data.get("gross",0), "refund":data.get("refund",0),
        "net_coll":data.get("net_coll",0), "deal_val":deal_val, "pcdv":pcdv_total,
        "cmr_sent":cmr_sent,"cmr_recd":cmr_recd,"cmr_pct":round(cmr_pct,1),
        "ss_sent":ss_sent,"ss_recd":ss_recd,"ss_pct":round(ss_pct,1),
        # Spot period deal values (KCD)
        "spot1_dv":fnt_dv.get("1_12",0), "spot2_dv":fnt_dv.get("20_30",0),
        "spot1_pcdv":pcdv_1_12,"spot2_pcdv":pcdv_20_30,
        # Spot period txns (CSD)
        "sp2_6_txn":spot_txn.get("2_6",0), "sp7_12_txn":spot_txn.get("7_12",0),
        "sp20_30_txn":spot_txn.get("20_30",0),
        "im_star_count":im_count, "im_star_pro_count":im_pro_count,
    }

    # ── NURSERY (L1, ageing ≤ 90 days) ────────────────────
    productivity = data.get("txns",0)  # total txns as proxy for productivity
    _fy_ce = emp.get("FY_Ageing", emp.get("Ageing",999))
    is_nursery = desig in ("L1","") and (_fy_ce <= 90 or emp.get("_nurs_bucket","60-90") == "60D")

    if is_nursery:
        min_prod = S["nursery_min_prod"]
        bucket    = emp.get("vintage_label", emp.get("_nurs_bucket","60-90"))
        fy_age_n  = emp.get("FY_Ageing", emp.get("Ageing", 999))
        # min_prod: 2 for 0-60D (from 2nd txn), 3 for 61-90D (from 3rd txn)
        if bucket == "60D" or fy_age_n <= 60:
            min_prod = 2   # first 60 days of TA tenure
        else:
            min_prod = S["nursery_min_prod"]  # 3 for days 61-90
        _cs_ok    = emp.get("Client_Status_Eligible", True)
        eligible  = _cs_ok and (productivity >= min_prod)
        inc = int(productivity * S["nursery_inc_per_txn"]) if eligible else 0
        mult = _nursery_mult(cmr_sent, cmr_recd, cmr_pct,
                             S["nursery_sent_max"], S["nursery_grid"],
                             S["nursery_table"], bucket=bucket,
                             grid_60d=None, fy_age=fy_age_n, vertical=vert)
        gross_inc = round(inc * mult, 0)
        out.update({
            "scheme":"Nursery", "productivity":productivity,
            "eligible":"Yes" if (_cs_ok and productivity>=min_prod) else "No",
            "renewal_mult":mult, "base_inc":inc, "gross_inc":gross_inc,
            "total_inc":int(gross_inc),
            "sp2_6_gross":0,"sp7_12_gross":0,"sp20_30_gross":0,
        })
        return out

    # ── TELE ANNUAL CSD ────────────────────────────────────
    if vert == "CSD":
        if desig == "L1":
            tgt_pcdv = S["csd_targets"].get(vint) or {"0-90":1300,"90-270":1800,"270+":2000}.get(vint, 1800)
            grid = _milestone(pcdv_total, tgt_pcdv, S["csd_milestones"])
            incr = round(max(0,pcdv_total-tgt_pcdv)*client_a*S["csd_incr_rate"]/1000)*1000 if grid>0 else 0
            base_incentive = grid + incr
            cmr_tgt = emp.get("CMR_Target_Pct", S["csd_cmr_tgt"])
            mult_val = _cmr_mult(cmr_pct, cmr_tgt, S["csd_cmr_mult"])
            gross_inc = round(base_incentive * mult_val, 0)
            # Spot (Exec-CSD)
            sp2_6   = _txn_spot(spot_txn.get("2_6",0),  S["csd_spot_2_6"])
            sp7_12  = _txn_spot(spot_txn.get("7_12",0), S["csd_spot_7_12"])
            _sp20_base = _txn_spot(spot_txn.get("20_30",0), S["csd_spot_20_30"])
            # Spot 20-30: SS+ receipts in that period × 2x bonus
            _ss_20_30 = data.get("ss_spot20_30", 0)
            sp20_30 = _sp20_base * (2 if _ss_20_30 > 0 else 1)
            total = int(gross_inc) + sp2_6 + sp7_12 + sp20_30
            out.update({
                "scheme":f"TA CSD Exec {vint}", "target_pcdv":tgt_pcdv,
                "incr_amt":int(incr), "incentive_grid":grid,
                "base_inc":int(base_incentive), "cmr_mult":mult_val,
                "gross_inc":int(gross_inc),
                "sp2_6_gross":sp2_6,"sp7_12_gross":sp7_12,"sp20_30_gross":sp20_30,
                "total_inc":total, "paid_inc":0, "bal_inc":total,
            })

        elif desig == "L2":
            # Rel'n Mgr CSD: team aggregate / HC
            tgt_pcdv = S["csd_targets"].get(vint) or {"0-90":1300,"90-270":1800,"270+":2000}.get(vint, 1800)
            grid = _milestone(pcdv_total, tgt_pcdv, S["csd_milestones"])
            incr = round(max(0,pcdv_total-tgt_pcdv)*client_a*S["csd_incr_rate"]/1000)*1000 if grid>0 else 0
            base_incentive = grid + incr
            mult_val = _cmr_mult(cmr_pct, emp.get("CMR_Target_Pct", S["csd_cmr_tgt"]), S["csd_cmr_mult"])
            gross_inc = round(base_incentive * mult_val, 0)
            # RM Spot: based on Productvity (= team txns / HC), not raw txn count
            sp2_6_prod  = spot_txn.get("2_6",0)  / max(hc,1)
            sp7_12_prod = spot_txn.get("7_12",0) / max(hc,1)
            sp20_30_prod= spot_txn.get("20_30",0)/ max(hc,1)
            # RM spot slab: threshold on Productvity
            # 2-6: prod>=1.0 → 1000, else 0
            # 7-12: prod>=1.5 → 1500, else 0
            # 20-30: prod>=3.0 → 1850, prod>=2.5 → 1550, else 0
            sp2_6   = 1000 if sp2_6_prod >= 1.0 else 0
            sp7_12  = 1500 if sp7_12_prod >= 1.5 else 0
            sp20_30 = (1850 if sp20_30_prod >= 3.0 else
                       1550 if sp20_30_prod >= 2.5 else 0)
            total = int(gross_inc) + sp2_6 + sp7_12 + sp20_30
            out.update({
                "scheme":f"TA CSD RM {vint}", "target_pcdv":tgt_pcdv,
                "incr_amt":int(incr),"incentive_grid":grid,
                "base_inc":int(base_incentive),"cmr_mult":mult_val,
                "gross_inc":int(gross_inc),
                "sp2_6_prod":round(sp2_6_prod,2),"sp7_12_prod":round(sp7_12_prod,2),
                "sp20_30_prod":round(sp20_30_prod,2),
                "sp2_6_gross":sp2_6,"sp7_12_gross":sp7_12,"sp20_30_gross":sp20_30,
                "total_inc":total,"paid_inc":0,"bal_inc":total,
            })

        elif desig == "L3":  # BM-CSD
            _gc=data.get("gross",0); _rv=data.get("refund",0); _nr=_gc-_rv
            # BM-CSD stores in Crores (matching sir's file)
            _net_cr=round(_nr/1e7,6); _coll_cr=round(_gc/1e7,6); _ref_cr=round(_rv/1e7,6)
            coll_target=emp.get("Coll_Target",0)  # in Crores from target file
            ach_pct=(_net_cr/coll_target*100) if coll_target>0 else 0
            incentive=_bm_milestone(ach_pct,S["bm_csd"]); total=int(incentive)
            out.update({
                "scheme":"BM CSD","gross_coll":_gc,"refund":_rv,"net_coll":_nr,
                "gross_coll_cr":_coll_cr,"refund_cr":_ref_cr,"net_coll_cr":_net_cr,
                "coll_target":coll_target,"ach_pct":round(ach_pct,2),
                "payout_eligible":"Yes" if incentive>0 else "No",
                "base_inc":incentive,"total_inc":total,"paid_inc":0,"bal_inc":total,
                "sp2_6_gross":0,"sp7_12_gross":0,"sp20_30_gross":0,
            })

        else:  # L4+ CH-CSD
            net_coll = data.get("net_coll",0)
            coll_target = emp.get("Coll_Target",0)
            ach_pct = (net_coll/coll_target*100) if coll_target>0 else 0
            incentive = _bm_milestone(ach_pct, S["ch_csd"])
            total = int(incentive)
            out.update({
                "scheme":"CH CSD","net_coll":net_coll,
                "coll_target":coll_target,"ach_pct":round(ach_pct,2),
                "base_inc":incentive,"total_inc":total,"paid_inc":0,"bal_inc":total,
                "sp2_6_gross":0,"sp7_12_gross":0,"sp20_30_gross":0,
            })

    # ── TELE ANNUAL KCD ────────────────────────────────────
    elif vert == "KCD":
        spot1_slabs = S["kcd_25cr_1_12"]  if is_25cr else S["kcd_spot_1_12"]
        spot2_slabs = S["kcd_25cr_20_30"] if is_25cr else S["kcd_spot_20_30"]
        rm1_slabs   = S["kcd_rm_1_12"]
        rm2_slabs   = S["kcd_rm_20_30"]
        # SS+ Multiplier for KCD: CMR+1 Ren% (Ren%.1) >= 2/3 → 1.0, else 0.5
        # (Confirmed from all 24 RM-KCD employees matching 100%)
        _cmr1_pct = cmr_prev.get("pct", 0) / 100
        ss_m = 1.0 if _cmr1_pct >= (2/3) else 0.5

        if desig == "L1":
            tgt_pcdv = S.get("kcd_target_map",{}).get(
                (str(emp.get("Group","KCD") or "KCD").strip(),vint)) or \
               {"KCD":{"0-270":7100,"270-2Yr":8100,"2Yr+":9100},"KCD-25cr":{"0-270":6480,"270-2Yr":9100,"2Yr+":10000}}.get(
                str(emp.get("Group","KCD") or "KCD").strip(),{}).get(vint,1800)
            grid   = _milestone(pcdv_total, tgt_pcdv, S["kcd_milestones"])
            incr   = round(max(0,pcdv_total-tgt_pcdv)*client_a*S.get("csd_incr_rate",0.03)/1000)*1000 if grid>0 else 0
            gross_inc = round((grid+incr)*ss_m,0)

            sp1_inc = _pcdv_spot(pcdv_1_12,  spot1_slabs)
            sp2_inc = _pcdv_spot(pcdv_20_30, spot2_slabs)
            sp1_gross = round(sp1_inc * ss_m + im_count*1000,0)   # spot 1-12
            sp2_gross = round(sp2_inc * ss_m + im_pro_count*1000,0) # spot 20-30 (+28-30 IM star pro)
            total = int(gross_inc) + int(sp1_gross) + int(sp2_gross)
            out.update({
                "scheme":f"TA KCD Exec {vint}","target_pcdv":tgt_pcdv,
                "incentive_grid":grid,"incr_amt":int(incr),
                "base_inc":int(grid+incr),"ss_mult":ss_m,"gross_inc":int(gross_inc),
                "spot1_inc":sp1_inc,"spot1_gross":int(sp1_gross),
                "spot2_inc":sp2_inc,"spot2_gross":int(sp2_gross),
                "total_inc":total,"paid_inc":0,"bal_inc":total,
                "sp2_6_gross":0,"sp7_12_gross":0,"sp20_30_gross":int(sp2_gross),
            })

        elif desig == "L2":
            tgt_pcdv = S.get("kcd_target_map",{}).get(
                (str(emp.get("Group","KCD") or "KCD").strip(),vint)) or \
               {"KCD":{"0-270":7100,"270-2Yr":8100,"2Yr+":9100},"KCD-25cr":{"0-270":6480,"270-2Yr":9100,"2Yr+":10000}}.get(
                str(emp.get("Group","KCD") or "KCD").strip(),{}).get(vint,1800)
            grid     = _milestone(pcdv_total,tgt_pcdv,S["kcd_milestones"])
            gross_inc= round(grid*ss_m,0)
            payout_elig = gross_inc > 0
            booster = 0  # configurable in future
            final_inc = int(gross_inc + booster)
            sp1_inc = _pcdv_spot(pcdv_1_12,  rm1_slabs)
            sp2_inc = _pcdv_spot(pcdv_20_30, rm2_slabs)
            sp1_gross = round(sp1_inc*ss_m + im_count*0.5*1000,0)   # team im count / team
            sp2_gross = round(sp2_inc*ss_m + im_pro_count*1000,0)
            im_ss_spot= im_pro_count * 1000  # 28-30 spot
            total = final_inc + int(sp1_gross) + int(sp2_gross) + im_ss_spot
            out.update({
                "scheme":f"TA KCD RM {vint}","target_pcdv":tgt_pcdv,
                "incentive_grid":grid,"payout_eligible":"Yes" if payout_elig else "No",
                "ss_mult":ss_m,"base_inc":int(gross_inc),"booster":booster,
                "final_inc":final_inc,
                "spot1_inc":sp1_inc,"spot1_gross":int(sp1_gross),
                "spot2_inc":sp2_inc,"spot2_gross":int(sp2_gross),
                "im_ss_spot":im_ss_spot,
                "total_inc":total,"paid_inc":0,"bal_inc":total,
                "sp2_6_gross":0,"sp7_12_gross":0,"sp20_30_gross":int(sp2_gross),
            })

        elif desig == "L3":  # BM-KCD
            _gc=data.get("gross",0); _rv=data.get("refund",0); _nr=_gc-_rv
            coll_target=emp.get("Coll_Target",0)  # KCD target in Rs
            ach_pct=(_nr/coll_target*100) if coll_target>0 else 0
            incentive=_bm_milestone(ach_pct,S["bm_kcd"]); total=int(incentive)
            out.update({
                "scheme":"BM KCD","gross_coll":_gc,"refund":_rv,"net_coll":_nr,
                "gross_coll_cr":_gc,"refund_cr":_rv,"net_coll_cr":_nr,  # KCD: all Rs (no conversion)
                "coll_target":coll_target,"ach_pct":round(ach_pct,2),
                "payout_eligible":"Yes" if incentive>0 else "No",
                "base_inc":incentive,"total_inc":total,"paid_inc":0,"bal_inc":total,
                "sp2_6_gross":0,"sp7_12_gross":0,"sp20_30_gross":0,
            })

        else:  # CH-KCD
            net_coll = data.get("net_coll",0)
            coll_target = emp.get("Coll_Target",0)
            ach_pct = (net_coll/coll_target*100) if coll_target>0 else 0
            incentive = _bm_milestone(ach_pct, S["ch_kcd"])
            total = int(incentive)
            out.update({
                "scheme":"CH KCD","net_coll":net_coll,
                "ach_pct":round(ach_pct,2),"base_inc":incentive,
                "total_inc":total,"paid_inc":0,"bal_inc":total,
                "sp2_6_gross":0,"sp7_12_gross":0,"sp20_30_gross":0,
            })

    return out


# ──────────────────────────────────────────────────────────
# BIG TICKET (from deal-level data in receipt)
# ──────────────────────────────────────────────────────────
def calc_big_ticket(rec, eid_str, desig, vert, slabs, is_l2=False):
    empty = {"3L+":0,"2L+":0,"1L+":0,"10L+":0,"8L+":0,"5L+":0}
    ec  = find_col(rec,["Sales Exec ID","EMP ID"])
    mgc = find_col(rec,["Old Sales HOD-3 ID","Manager Id"])
    upc = find_col(rec,["Unique","Upsell"])
    if not ec: return 0, empty

    if is_l2 and mgc:
        mask = rec[mgc].astype(str).str.split(".").str[0].str.strip()==eid_str
    else:
        mask = rec[ec].astype(str).str.split(".").str[0].str.strip()==eid_str
    r = rec[mask].copy()
    if len(r) == 0: return 0, empty

    # Only IM Star/Leader type deals for big ticket
    if upc and upc in r.columns:
        im_mask = r[upc].fillna("").astype(str).str.upper().apply(
            lambda x: any(k in x for k in IM_STAR_LEADER_KW))
        r = r[im_mask]
    if len(r) == 0: return 0, empty

    # Re-locate deal value column on the filtered subset to avoid KeyError
    dvc = find_col(r, ["Deal Val (WT)","Deal Value","Deal Val","WT AMT","WT_AMT"])
    if not dvc: return 0, empty

    try:
        deal_vals_lakh = pd.to_numeric(r[dvc], errors="coerce").fillna(0) / 100000
    except (KeyError, TypeError):
        return 0, empty

    counts = dict(empty)
    total_bt = 0
    for dv in deal_vals_lakh:
        for s in slabs:
            if dv >= s.get("Min_Lakh",0):
                size_key = s.get("Deal_Size","")
                counts[size_key] = counts.get(size_key,0) + 1
                total_bt += s.get("Per_Deal",0)
                break
    return total_bt, counts



# ──────────────────────────────────────────────────────────
# MILESTONE TARGET FILE LOADER
# ──────────────────────────────────────────────────────────
def load_ta_targets(f):
    """
    Load milestone targets from the uploaded target file.
    Expected sheets:
      L4       : L4 Name, Target (collection target in Cr)
      L3       : Employee Name, Target (collection target in Cr)
      L1 Renewal: IIL (Emp ID), Employee Name, TGT CMR 100% (decimal)
      L2 Target : IIL (Emp ID), Employee Name, % TGT (decimal)
    Returns dict with keys: l1_cmr, l2_cmr, l3_coll, l4_coll
      l1_cmr  : {emp_id_str: cmr_target_pct (0-100)}
      l2_cmr  : {emp_id_str: cmr_target_pct (0-100)}
      l3_coll : {emp_name_lower: coll_target_rs}
      l4_coll : {l4_name_lower:  coll_target_rs}
    """
    empty = {"l1_cmr":{}, "l2_cmr":{}, "l3_coll":{}, "l4_coll":{}}
    if f is None: return empty
    try:
        _tb = f.getvalue() if hasattr(f,"getvalue") else (open(f,"rb").read() if isinstance(f,str) else f.read())
        xl = pd.ExcelFile(_io.BytesIO(_tb))
        norms = {s.strip().upper(): s for s in xl.sheet_names}

        # L1 Renewal CMR targets
        l1_cmr = {}
        sh1 = norms.get("L1 RENEWAL") or norms.get("L1RENEWAL") or norms.get("L1")
        if sh1:
            df1 = pd.read_excel(_io.BytesIO(_tb), sheet_name=sh1)
            df1.columns = [str(c).strip() for c in df1.columns]
            iil_c = find_col(df1, ["IIL","Employee ID","Emp ID","EmpID"])
            tgt_c = find_col(df1, ["TGT CMR 100%","CMR Target","Target CMR","TGT CMR","CMR%","Target"])
            if iil_c and tgt_c:
                for _, row in df1.iterrows():
                    eid = str(row[iil_c]).strip().split(".")[0]
                    if not eid or eid.lower() in ("nan","none",""): continue
                    val = _sf(row[tgt_c])
                    # Convert decimal (0.4) → percentage (40.0)
                    l1_cmr[eid] = round(val * 100 if val <= 1.0 else val, 4)

        # L2 CMR targets
        l2_cmr = {}
        sh2 = norms.get("L2 TARGET") or norms.get("L2TARGET") or norms.get("L2")
        if sh2:
            df2 = pd.read_excel(_io.BytesIO(_tb), sheet_name=sh2)
            df2.columns = [str(c).strip() for c in df2.columns]
            iil_c = find_col(df2, ["IIL","Employee ID","Emp ID"])
            tgt_c = find_col(df2, ["% TGT","TGT","Target","CMR Target","CMR%"])
            if iil_c and tgt_c:
                for _, row in df2.iterrows():
                    eid = str(row[iil_c]).strip().split(".")[0]
                    if not eid or eid.lower() in ("nan","none",""): continue
                    val = _sf(row[tgt_c])
                    l2_cmr[eid] = round(val * 100 if val <= 1.0 else val, 4)

        # L3 collection targets (in Crores → convert to Rs.)
        l3_coll = {}
        sh3 = norms.get("L3")
        if sh3:
            df3 = pd.read_excel(_io.BytesIO(_tb), sheet_name=sh3)
            df3.columns = [str(c).strip() for c in df3.columns]
            name_c = find_col(df3, ["Employee Name","Name","L3 Name","L3Name"])
            tgt_c  = find_col(df3, ["Target","Collection Target","Coll Target"])
            eid_c  = find_col(df3, ["Employee ID","EID","IIL","Emp ID"])
            if name_c and tgt_c:
                for _, row in df3.iterrows():
                    nm = str(row[name_c]).strip()
                    if not nm or nm.lower() in ("nan","none",""): continue
                    val = _sf(row[tgt_c])
                    # BM-CSD target stored in Crores (< 1000) — keep as-is
                    l3_coll[nm.lower()] = val
                    if eid_c:
                        eid_val = str(row[eid_c]).strip().split('.')[0]
                        if eid_val and eid_val not in ("nan","none",""):
                            l3_coll[eid_val] = val

        # L4 collection targets
        l4_coll = {}
        sh4 = norms.get("L4")
        if sh4:
            df4 = pd.read_excel(_io.BytesIO(_tb), sheet_name=sh4)
            df4.columns = [str(c).strip() for c in df4.columns]
            name_c = find_col(df4, ["L4 Name","Name","Employee Name"])
            tgt_c  = find_col(df4, ["Target","Collection Target","Coll Target"])
            if name_c and tgt_c:
                for _, row in df4.iterrows():
                    nm = str(row[name_c]).strip()
                    if not nm or nm.lower() in ("nan","none",""): continue
                    val = _sf(row[tgt_c])
                    l4_coll[nm.lower()] = val  # Crore

        return {"l1_cmr": l1_cmr, "l2_cmr": l2_cmr,
                "l3_coll": l3_coll, "l4_coll": l4_coll}
    except Exception as e:
        st.warning(f"⚠️ Could not load target file: {e}")
        return empty

# ──────────────────────────────────────────────────────────
# OUTPUT SHEET BUILDER
# ──────────────────────────────────────────────────────────
def write_sheet(w, df, sheet_name, header_fmt, note="", merged_headers=None):
    """Write a DataFrame to sheet with optional merged header row."""
    start = 2 if merged_headers else 1
    df.to_excel(w, sheet_name=sheet_name, index=False, startrow=start)
    ws = w.sheets[sheet_name]
    for ci, col in enumerate(df.columns):
        ws.write(start, ci, col, header_fmt)
        ws.set_column(ci, ci, max(14, len(str(col))+2))
    if merged_headers:
        for (merge_start, merge_end, label) in merged_headers:
            ws.merge_range(0, merge_start, 0, merge_end, label, header_fmt)
    if note:
        ws.write(0 if not merged_headers else 1, 0, note, None)
    ws.freeze_panes(start+1, 0)


def build_excel_output(results_dict, sel_month):
    """Build final Excel with all sheets matching sirs file structure."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
        wb  = w.book
        hdr = wb.add_format({"bold":True,"bg_color":"#1F4E79","font_color":"#FFFFFF","border":1,"font_size":9})
        grn = wb.add_format({"bold":True,"bg_color":"#375623","font_color":"#FFFFFF","border":1,"font_size":9})
        org = wb.add_format({"bold":True,"bg_color":"#843C0C","font_color":"#FFFFFF","border":1,"font_size":9})
        pur = wb.add_format({"bold":True,"bg_color":"#7030A0","font_color":"#FFFFFF","border":1,"font_size":9})
        gry = wb.add_format({"bold":True,"bg_color":"#595959","font_color":"#FFFFFF","border":1,"font_size":9})

        hierarchy_l1  = ["Employee ID","Employee Name","L2 ID","L2 Name","L3 ID","L3 Name","L4 ID","L4 Name","L5 ID","L5 Name","L6 ID","L6 Name"]
        hierarchy_bm  = ["Employee ID","Employee Name","L3 ID","L3 Name","L4 ID","L4 Name","L5 ID","L5 Name","L6 ID","L6 Name","HC","L2"]
        hierarchy_ch  = ["Employee ID","Employee Name","L5 ID","L5 Name","L6 ID","L6 Name"]

        def _dfcols(rows, cols):
            d = pd.DataFrame(rows)
            out = {}
            for c in cols:
                out[c] = d[c].values if c in d.columns else [""]*len(d)
            return pd.DataFrame(out)

        def _getv(r, *keys):
            for k in keys:
                if k in r and r[k] not in (None,""): return r[k]
            return ""

        # ── Nursery ────────────────────────────────────────
        nursery_rows = results_dict.get("nursery",[])
        if nursery_rows:
            cols_n = ["Employee ID","Employee Name","L2 ID","L2 Name","L3 ID","L3 Name",
                      "L4 ID","L4 Name","L5 ID","L5 Name","L6  ID","L6 Name",
                      "Location","Client-A","Client-C",
                      "MDC","Star","Leader","WS-M","WS-A","IVE",
                      "MDC.1","Star.1","Leader.1","WS-M.1","WS-A.1","IVE.1",
                      "Email Id","Joining Date",
                      "IIL Vertical Name",
                      "CSD to KCD\nMovement/New Joining\n<=90D",
                      "60D/90D","Productvity","Eligible",
                      "Sent","Recd","Ren %",
                      "Renewal\nMultiplier","Incentive","Gross Incentive",
                      "Paid\nIncentive","Balance\nIncentive"]
            rows_n = []
            for r in nursery_rows:
                rows_n.append({
                    "Employee ID":r.get("eid",""),"Employee Name":r.get("name",""),
                    "L2 ID":r.get("L2 ID",""),"L2 Name":r.get("L2 Name",""),
                    "L3 ID":r.get("L3 ID",""),"L3 Name":r.get("L3 Name",""),
                    "L4 ID":r.get("L4 ID",""),"L4 Name":r.get("L4 Name",""),
                    "L5 ID":r.get("L5 ID",""),"L5 Name":r.get("L5 Name",""),
                    "L6  ID":r.get("L6 ID",""),"L6 Name":r.get("L6 Name",""),
                    "Location":r.get("Location",""),"Client-A":r.get("Client-A",0),
                    "Client-C":r.get("Client-C",0),
                    "MDC":r.get("MDC",0),"Star":r.get("Star",0),"Leader":r.get("Leader",0),
                    "WS-M":r.get("WS-M",0),"WS-A":r.get("WS-A",0),"IVE":r.get("IVE",0),
                    "MDC.1":r.get("MDC.1",0),"Star.1":r.get("Star.1",0),"Leader.1":r.get("Leader.1",0),
                    "WS-M.1":r.get("WS-M.1",0),"WS-A.1":r.get("WS-A.1",0),"IVE.1":r.get("IVE.1",0),
                    "Email Id":r.get("Email Id",""),
                    "Joining Date":str(r.get("Joining Date",""))[:10],
                    "IIL Vertical Name":"Tele Annual "+r.get("Vertical","CSD"),
                    "CSD to KCD\nMovement/New Joining\n<=90D":"L1",
                    "60D/90D":r.get("vintage_label","60-90"),
                    "Productvity":r.get("productivity",0),
                    "Eligible":r.get("eligible","No"),
                    "Sent":r.get("cmr_sent",0),"Recd":r.get("cmr_recd",0),
                    "Ren %":round(r.get("cmr_pct",0)/100,4),
                    "Renewal\nMultiplier":r.get("renewal_mult",0),
                    "Incentive":r.get("base_inc",0),
                    "Gross Incentive":r.get("gross_inc",0),
                    "Paid\nIncentive":0,"Balance\nIncentive":r.get("gross_inc",0),
                })
            df_n = pd.DataFrame(rows_n).reindex(columns=cols_n, fill_value="")

            df_n.to_excel(w, sheet_name="Nursery Scheme", index=False, startrow=1)
            ws = w.sheets["Nursery Scheme"]
            for ci,col in enumerate(cols_n): ws.write(1,ci,col,pur)
            ws.set_column(0,len(cols_n)-1,16)
            ws.freeze_panes(2,0)

        # ── Exec-CSD ───────────────────────────────────────
        csd_l1 = results_dict.get("exec_csd",[])
        if csd_l1:
            cols_e = ["Employee ID","Employee Name","L2 ID","L2 Name","L3 ID","L3 Name",
                      "L4 ID","L4 Name","L5 ID","L5 Name","L6 ID","L6 Name",
                      "Joining Date/Movement Date","IIL Vertical Name",
                      "Aeging","Vintage","Client-A","Client-C",
                      "Deal Value","PCDV","Target","Incremental\nPCDV\nAmt.",
                      "Sent","Recd","Ren %","Multiplier","Renewal\nTarget",
                      "Sent.1","Recd.1","Ren %.1",
                      "Incentive Grid","Incentive","Gross\nIncentive","Paid\nIncentive","Balance\nIncentive",
                      "Transaction","Gross\nIncentive","Paid\nIncentive","Balance\nIncentive",  # Spot 2-6
                      "Transaction","Gross\nIncentive","Paid\nIncentive","Balance\nIncentive",  # Spot 7-12
                      "Transaction","Gross\nIncentive","Paid\nIncentive","Balance\nIncentive",  # Spot 20-30
                      ]
            # Use unique column names for the repeated spot cols
            cols_out = ["Employee ID","Employee Name","L2 ID","L2 Name","L3 ID","L3 Name",
                        "L4 ID","L4 Name","L5 ID","L5 Name","L6  ID","L6 Name",
                        "Joining Date/Movement Date","IIL Vertical Name","Aeging","Vintage",
                        "Client-A","Client-C","Deal Value","PCDV","Target","Incremental\nPCDV\nAmt.",
                        "Sent","Recd","Ren %","Multiplier","Renewal\nTarget",
                        "Sent.1","Recd.1","Ren %.1",
                        "Incentive Grid","Incentive","Gross\nIncentive","Paid\nIncentive","Balance\nIncentive",
                        "Transaction","Gross\nIncentive.1","Paid\nIncentive.1","Balance\nIncentive.1",
                        "Transaction.1","Gross\nIncentive.2","Paid\nIncentive.2","Balance\nIncentive.2",
                        "Transaction.2","Gross\nIncentive.3","Paid\nIncentive.3","Balance\nIncentive.3"]
            rows_e = []
            for r in csd_l1:
                rows_e.append({
                    "Employee ID":r["eid"],"Employee Name":r["name"],
                    "L2 ID":r.get("L2 ID",""),"L2 Name":r.get("L2 Name",""),
                    "L3 ID":r.get("L3 ID",""),"L3 Name":r.get("L3 Name",""),
                    "L4 ID":r.get("L4 ID",""),"L4 Name":r.get("L4 Name",""),
                    "L5 ID":r.get("L5 ID",""),"L5 Name":r.get("L5 Name",""),
                    "L6  ID":r.get("L6 ID",""),"L6 Name":r.get("L6 Name",""),
                    "Joining Date/Movement Date":str(r.get("Joining Date",""))[:10],
                    "IIL Vertical Name":"Tele Annual CSD",
                    "Aeging":r.get("Ageing",0),"Vintage":r.get("Vintage",""),
                    "Client-A":r.get("Client-A_Agg",r.get("Client-A",0)),
                    "Client-C":r.get("Client-C_Agg",r.get("Client-C",0)),
                    "Deal Value":round(r.get("deal_val",0),2),"PCDV":round(r.get("pcdv",0),2),
                    "Target":r.get("target_pcdv",0),"Incremental\nPCDV\nAmt.":r.get("incr_amt",0),
                    "Sent":r.get("cmr_sent",0),"Recd":r.get("cmr_recd",0),
                    "Ren %":round(r.get("cmr_pct",0)/100,4),
                    "Multiplier":r.get("cmr_mult",0),"Renewal\nTarget":r.get("cmr_target_pct", r.get("csd_cmr_tgt",40))/100,
                    "Sent.1":r.get("cmr1_sent",0),"Recd.1":r.get("cmr1_recd",0),"Ren %.1":round(r.get("cmr1_pct",0)/100,4),
                    "Incentive Grid":r.get("incentive_grid",0),"Incentive":r.get("base_inc",0),
                    "Gross\nIncentive":r.get("gross_inc",0),"Paid\nIncentive":0,
                    "Balance\nIncentive":r.get("gross_inc",0),
                    "Transaction":r.get("sp2_6_txn",0),"Gross\nIncentive.1":r.get("sp2_6_gross",0),"Paid\nIncentive.1":0,"Balance\nIncentive.1":r.get("sp2_6_gross",0),
                    "Transaction.1":r.get("sp7_12_txn",0),"Gross\nIncentive.2":r.get("sp7_12_gross",0),"Paid\nIncentive.2":0,"Balance\nIncentive.2":r.get("sp7_12_gross",0),
                    "Transaction.2":r.get("sp20_30_txn",0),"Gross\nIncentive.3":r.get("sp20_30_gross",0),"Paid\nIncentive.3":0,"Balance\nIncentive.3":r.get("sp20_30_gross",0),
                })
            df_e = pd.DataFrame(rows_e).reindex(columns=cols_out, fill_value="")
            df_e.to_excel(w, sheet_name="Exec-CSD", index=False, startrow=2)
            ws = w.sheets["Exec-CSD"]
            # Row 0: merged group headers
            ws.merge_range(0,35,0,38,"Spot 2nd to 6th",hdr)
            ws.merge_range(0,39,0,42,"Spot 7th to 12th",hdr)
            ws.merge_range(0,43,0,46,"Spot 20th to 30th",hdr)
            for ci,col in enumerate(cols_out): ws.write(1,ci,col,grn)
            ws.set_column(0,len(cols_out)-1,14); ws.freeze_panes(3,0)

        # ── Rel'n Mgr-CSD ──────────────────────────────────
        csd_l2 = results_dict.get("rm_csd",[])
        if csd_l2:
            cols_rm = ["Employee ID","Employee Name","L2 ID","L2 Name","L3 ID","L3 Name",
                       "L4 ID","L4 Name","L5 ID","L5 Name","L6  ID","L6 Name",
                       "Joining Date","IIL Vertical Name","Group","<90\nVintage\nExec","HC",
                       "Client-A","Client-C","Deal Value","PCDV",
                       "Sent","Recd","Ren %","Renewal\nTarget",
                       "Sent.1","Recd.1","Ren %.1",
                       "Incremental\nPCDV\nAmt.","Incentive\nGrid","Incentive","Gross Incentive","Paid\nIncentive","Balance\nIncentive",
                       "Transaction","Productvity","Gross\nIncentive.1","Paid\nIncentive.1","Balance\nIncentive.1",
                       "Transaction.1","Productvity.1","Gross\nIncentive.2","Paid\nIncentive.2","Balance\nIncentive.2",
                       "Transaction.2","Productvity.2","Gross\nIncentive.3","Paid\nIncentive.3","Balance\nIncentive.3"]
            rows_rm = []
            for r in csd_l2:
                rows_rm.append({
                    "Employee ID":r["eid"],"Employee Name":r["name"],
                    "L2 ID":r.get("L2 ID",""),"L2 Name":r.get("L2 Name",""),
                    "L3 ID":r.get("L3 ID",""),"L3 Name":r.get("L3 Name",""),
                    "L4 ID":r.get("L4 ID",""),"L4 Name":r.get("L4 Name",""),
                    "L5 ID":r.get("L5 ID",""),"L5 Name":r.get("L5 Name",""),
                    "L6  ID":r.get("L6 ID",""),"L6 Name":r.get("L6 Name",""),
                    "Joining Date":str(r.get("Joining Date",""))[:10],
                    "IIL Vertical Name":"Tele Annual CSD","Group":"",
                    "<90\nVintage\nExec":r.get("lt90_count",0),"HC":r.get("HC",0),
                    "Client-A":r.get("Client-A_Agg",r.get("Client-A",0)),"Client-C":r.get("Client-C_Agg",r.get("Client-C",0)),
                    "Deal Value":round(r.get("deal_val",0),2),"PCDV":round(r.get("pcdv",0),2),
                    "Sent":r.get("cmr_sent",0),"Recd":r.get("cmr_recd",0),
                    "Ren %":round(r.get("cmr_pct",0)/100,4),
                    "Renewal\nTarget":r.get("cmr_target_pct", r.get("csd_cmr_tgt",40))/100,
                    "Sent.1":r.get("cmr1_sent",0),"Recd.1":r.get("cmr1_recd",0),"Ren %.1":round(r.get("cmr1_pct",0)/100,4),
                    "Incremental\nPCDV\nAmt.":r.get("incr_amt",0),"Incentive\nGrid":r.get("incentive_grid",0),
                    "Incentive":r.get("base_inc",0),"Gross Incentive":r.get("gross_inc",0),
                    "Paid\nIncentive":0,"Balance\nIncentive":r.get("gross_inc",0),
                    "Transaction":r.get("sp2_6_txn",0),
                    "Productvity":round(r.get("sp2_6_txn",0)/max(r.get("HC",1),1),4),
                    "Gross\nIncentive.1":r.get("sp2_6_gross",0),"Paid\nIncentive.1":0,"Balance\nIncentive.1":r.get("sp2_6_gross",0),
                    "Transaction.1":r.get("sp7_12_txn",0),
                    "Productvity.1":round(r.get("sp7_12_txn",0)/max(r.get("HC",1),1),2),
                    "Gross\nIncentive.2":r.get("sp7_12_gross",0),"Paid\nIncentive.2":0,"Balance\nIncentive.2":r.get("sp7_12_gross",0),
                    "Transaction.2":r.get("sp20_30_txn",0),
                    "Productvity.2":round(r.get("sp20_30_txn",0)/max(r.get("HC",1),1),2),
                    "Gross\nIncentive.3":r.get("sp20_30_gross",0),"Paid\nIncentive.3":0,"Balance\nIncentive.3":r.get("sp20_30_gross",0),
                })
            df_rm = pd.DataFrame(rows_rm).reindex(columns=cols_rm, fill_value="")
            df_rm.to_excel(w, sheet_name="Rel'n Mgr-CSD", index=False, startrow=2)
            ws = w.sheets["Rel'n Mgr-CSD"]
            ws.merge_range(0,34,0,38,"Spot 2nd to 6th",hdr)
            ws.merge_range(0,39,0,43,"Spot 7th to 12th",hdr)
            ws.merge_range(0,44,0,48,"Spot 20th to 30th",hdr)
            for ci,col in enumerate(cols_rm): ws.write(1,ci,col,grn)
            ws.set_column(0,len(cols_rm)-1,14); ws.freeze_panes(3,0)

        # ── BM-CSD ─────────────────────────────────────────
        bm_csd_rows = results_dict.get("bm_csd",[])
        if bm_csd_rows:
            cols_bm = ["Employee ID","Employee Name","L3 ID","L3 Name","L4 ID","L4 Name",
                       "L5 ID","L5 Name","L6  ID","L6 Name","HC","L2",
                       "Joining Date","IIL Vertical Name","Location",
                       "Client-A","Client-C","Deal Value","PCDV","Collection","Refund","Net Collection",
                       "Target","Ach\n%","CMR\nSent","CMR\nReceived","Ren %",
                       "Payout\nEligibile","Incentive","Total Incentive","Gross Incentive",
                       "Paid\nIncentive","Balance\nIncentive"]
            rows_bm = []
            for r in bm_csd_rows:
                rows_bm.append({
                    "Employee ID":r["eid"],"Employee Name":r["name"],
                    "L3 ID":r.get("L3 ID",""),"L3 Name":r.get("L3 Name",""),
                    "L4 ID":r.get("L4 ID",""),"L4 Name":r.get("L4 Name",""),
                    "L5 ID":r.get("L5 ID",""),"L5 Name":r.get("L5 Name",""),
                    "L6  ID":r.get("L6 ID",""),"L6 Name":r.get("L6 Name",""),
                    "HC":r.get("HC",0),"L2":r.get("L2_Count",0),
                    "Joining Date":str(r.get("Joining Date",""))[:10],
                    "IIL Vertical Name":"Tele Annual CSD","Location":r.get("Location",""),
                    "Client-A":r.get("Client-A_Agg",r.get("Client-A",0)),"Client-C":r.get("Client-C_Agg",r.get("Client-C",0)),
                    "Deal Value":round(r.get("deal_val",0),2),"PCDV":round(r.get("pcdv",0),2),
                    "Collection":round(r.get("gross_coll",0),2),"Refund":round(r.get("refund",0),2),
                    "Net Collection":round(r.get("net_coll",0)/1e7,6),  # Crore
                    "Target":r.get("coll_target") or None,"Ach\n%":round(r.get("ach_pct",0)/100,4),
                    "CMR\nSent":r.get("cmr_sent",0),"CMR\nReceived":r.get("cmr_recd",0),
                    "Ren %":round(r.get("cmr_pct",0)/100,4),
                    "Payout\nEligibile":r.get("payout_eligible","No"),
                    "Incentive":r.get("base_inc",0),"Total Incentive":r.get("total_inc",0),
                    "Gross Incentive":r.get("total_inc",0),
                    "Paid\nIncentive":0,"Balance\nIncentive":r.get("total_inc",0),
                })
            df_bm = pd.DataFrame(rows_bm).reindex(columns=cols_bm, fill_value="")
            df_bm.to_excel(w, sheet_name="BM-CSD", index=False, startrow=1)
            ws = w.sheets["BM-CSD"]
            for ci,col in enumerate(cols_bm): ws.write(1,ci,col,grn)
            ws.set_column(0,len(cols_bm)-1,14); ws.freeze_panes(2,0)

        # ── BM-CSD Big Ticket ──────────────────────────────
        bt_bm_csd_rows = results_dict.get("bt_bm_csd",[])
        if bt_bm_csd_rows:
            cols_bt = ["Employee ID","Employee Name","L4 ID","L4 Name","L5 ID","L5 Name",
                       "L6  ID","L6 Name","Location","Joining Date","IIL Vertical Name",
                       "Net Collection","AOP","%",
                       "Sent","Recd","Ren %",
                       "3L+","2L+","1L+","Total","Eligible",
                       "Gross\nIncentive","Paid\nIncentive","Balance\nIncentive"]
            rows_bt = []
            for r in bt_bm_csd_rows:
                cnt = r.get("bt_counts",{})
                total_deals = sum(cnt.get(k,0) for k in ["3L+","2L+","1L+"])
                rows_bt.append({
                    "Employee ID":r["eid"],"Employee Name":r["name"],
                    "L4 ID":r.get("L4 ID",""),"L4 Name":r.get("L4 Name",""),
                    "L5 ID":r.get("L5 ID",""),"L5 Name":r.get("L5 Name",""),
                    "L6  ID":r.get("L6 ID",""),"L6 Name":r.get("L6 Name",""),
                    "Location":r.get("Location",""),
                    "Joining Date":str(r.get("Joining Date",""))[:10],
                    "IIL Vertical Name":"Tele Annual CSD",
                    "Net Collection":round(r.get("net_coll",0),2),
                    "AOP":r.get("coll_target",None),"%":round(r.get("net_coll_cr",r.get("net_coll",0)/1e7)/max(r.get("coll_target",0) or 1,1)*100,4),
                    "Sent":r.get("cmr_sent",0),"Recd":r.get("cmr_recd",0),
                    "Ren %":round(r.get("cmr_pct",0)/100,4),
                    "3L+":cnt.get("3L+",0),"2L+":cnt.get("2L+",0),"1L+":cnt.get("1L+",0),
                    "Total":total_deals,"Eligible":r.get("bt_eligible","NO"),
                    "Gross\nIncentive":r.get("bt_inc",0),"Paid\nIncentive":0,"Balance\nIncentive":r.get("bt_inc",0),
                })
            df_bt = pd.DataFrame(rows_bt).reindex(columns=cols_bt, fill_value="")
            df_bt.to_excel(w, sheet_name="BM-CSD Big Ticket", index=False, startrow=1)
            ws = w.sheets["BM-CSD Big Ticket"]
            for ci,col in enumerate(cols_bt): ws.write(1,ci,col,grn)
            ws.set_column(0,len(cols_bt)-1,14); ws.freeze_panes(2,0)

        # ── CH-CSD Big Ticket ──────────────────────────────
        bt_ch_csd_rows = results_dict.get("bt_ch_csd",[])
        if bt_ch_csd_rows:
            cols_btch = ["Employee ID","Employee Name","L5 ID","L5 Name","L6  ID","L6 Name",
                         "Location","Joining Date","IIL Vertical Name",
                         "Client-A","Client-C","Target\nOf 1L+ Deal",
                         "Net Collection","AOP","AOP%",
                         "Sent","Recd","Ren %",
                         "3L+","2L+","1L+","Total","Eligible",
                         "Gross\nIncentive","Paid\nIncentive","Balance\nIncentive"]
            rows_btch = []
            for r in bt_ch_csd_rows:
                cnt = r.get("bt_counts",{})
                total_deals = sum(cnt.get(k,0) for k in ["3L+","2L+","1L+"])
                rows_btch.append({
                    "Employee ID":r["eid"],"Employee Name":r["name"],
                    "L5 ID":r.get("L5 ID",""),"L5 Name":r.get("L5 Name",""),
                    "L6  ID":r.get("L6 ID",""),"L6 Name":r.get("L6 Name",""),
                    "Location":r.get("Location",""),
                    "Joining Date":str(r.get("Joining Date",""))[:10],
                    "IIL Vertical Name":"Tele Annual CSD",
                    "Client-A":r.get("Client-A",0),"Client-C":r.get("Client-C",0),
                    "Target\nOf 1L+ Deal":r.get("deal_target",0),
                    "Net Collection":round(r.get("net_coll",0),2),
                    "AOP":r.get("aop",0),"AOP%":round(r.get("aop_pct",0),2),
                    "Sent":r.get("cmr_sent",0),"Recd":r.get("cmr_recd",0),
                    "Ren %":round(r.get("cmr_pct",0)/100,4),
                    "3L+":cnt.get("3L+",0),"2L+":cnt.get("2L+",0),"1L+":cnt.get("1L+",0),
                    "Total":total_deals,"Eligible":r.get("bt_eligible","No"),
                    "Gross\nIncentive":r.get("bt_inc",0),"Paid\nIncentive":0,"Balance\nIncentive":r.get("bt_inc",0),
                })
            df_btch = pd.DataFrame(rows_btch).reindex(columns=cols_btch, fill_value="")
            df_btch.to_excel(w, sheet_name="CH-CSD Big Ticket", index=False, startrow=1)
            ws = w.sheets["CH-CSD Big Ticket"]
            for ci,col in enumerate(cols_btch): ws.write(1,ci,col,grn)
            ws.set_column(0,len(cols_btch)-1,14); ws.freeze_panes(2,0)

        # ── Exec-KCD ───────────────────────────────────────
        kcd_l1 = results_dict.get("exec_kcd",[])
        if kcd_l1:
            cols_kl1 = ["Employee ID","Employee Name","L2 ID","L2 Name","L3 ID","L3 Name",
                        "L4 ID","L4 Name","L5 ID","L5 Name","L6  ID","L6 Name",
                        "Joining Date/Movement Date","IIL Vertical Name","Group","Aeging","Vintage",
                        "Target","Client-A","Client-C","Deal Value","PCDV",
                        "Sent","Recd","Ren %","Sent.1","Recd.1","Ren %.1",
                        "SS+\nMultiplier","Target\nPCDV%","Incentive","Mulitipler\nPayout","Final\nIncentive",
                        "Gross Incentive","Paid\nIncentive","Balance\nIncentive",
                        "Deal Value.1","PCDV.1","Incentive.1","IM Star/Leader\nNew Sale","Gross Incentive.1","Paid\nIncentive.1","Balance\nIncentive.1",
                        "Deal Value.2","PCDV.2","Incentive.2","IM Star/Leader\nNew Sale.1","Gross Incentive.2","Paid\nIncentive.2","Balance\nIncentive.2"]
            rows_kl1=[]
            for r in kcd_l1:
                tgt = r.get("target_pcdv",0)
                pcdv = r.get("pcdv",0)
                tgt_pct = round(pcdv/tgt*100,2) if tgt>0 else 0
                rows_kl1.append({
                    "Employee ID":r["eid"],"Employee Name":r["name"],
                    "L2 ID":r.get("L2 ID",""),"L2 Name":r.get("L2 Name",""),
                    "L3 ID":r.get("L3 ID",""),"L3 Name":r.get("L3 Name",""),
                    "L4 ID":r.get("L4 ID",""),"L4 Name":r.get("L4 Name",""),
                    "L5 ID":r.get("L5 ID",""),"L5 Name":r.get("L5 Name",""),
                    "L6  ID":r.get("L6 ID",""),"L6 Name":r.get("L6 Name",""),
                    "Joining Date/Movement Date":str(r.get("Joining Date",""))[:10],
                    "IIL Vertical Name":"Tele Annual KCD","Group":r.get("Group",""),
                    "Aeging":r.get("Ageing",0),"Vintage":r.get("Vintage",""),
                    "Target":tgt,
                    "Client-A":r.get("Client-A",0),
                    "Client-C":r.get("Client-C",0),
                    "Deal Value":round(r.get("deal_val",0),2),"PCDV":round(pcdv,2),
                    "Sent":r.get("cmr_sent",0),"Recd":r.get("cmr_recd",0),
                    "Ren %":round(r.get("cmr_pct",0)/100,4),
                    "Sent.1":r.get("ss_sent",0),"Recd.1":r.get("ss_recd",0),
                    "Ren %.1":round(r.get("ss_pct",0)/100,4),
                    "SS+\nMultiplier":r.get("ss_mult",1),"Target\nPCDV%":tgt_pct,
                    "Incentive":r.get("incentive_grid",0),
                    "Mulitipler\nPayout":round(r.get("incentive_grid",0)*r.get("ss_mult",1),0),
                    "Final\nIncentive":r.get("gross_inc",0),
                    "Gross Incentive":r.get("gross_inc",0),"Paid\nIncentive":0,"Balance\nIncentive":r.get("gross_inc",0),
                    "Deal Value.1":round(r.get("fnt_dv",{}).get("1_12",0),2),
                    "PCDV.1":round(r.get("pcdv_1_12",0),2),
                    "Incentive.1":r.get("spot1_inc",0),"IM Star/Leader\nNew Sale":r.get("im_count",0),
                    "Gross Incentive.1":r.get("spot1_gross",0),"Paid\nIncentive.1":0,"Balance\nIncentive.1":r.get("spot1_gross",0),
                    "Deal Value.2":round(r.get("fnt_dv",{}).get("20_30",0),2),
                    "PCDV.2":round(r.get("pcdv_20_30",0),2),
                    "Incentive.2":r.get("spot2_inc",0),"IM Star/Leader\nNew Sale.1":r.get("im_pro_count",0),
                    "Gross Incentive.2":r.get("spot2_gross",0),"Paid\nIncentive.2":0,"Balance\nIncentive.2":r.get("spot2_gross",0),
                })
            df_kl1 = pd.DataFrame(rows_kl1).reindex(columns=cols_kl1, fill_value="")
            df_kl1.to_excel(w, sheet_name="Exec-KCD", index=False, startrow=2)
            ws = w.sheets["Exec-KCD"]
            ws.merge_range(0,36,0,42,"Spot 1st to 12th",org)
            ws.merge_range(0,43,0,49,"Spot 20th to 30th",org)
            for ci,col in enumerate(cols_kl1): ws.write(1,ci,col,org)
            ws.set_column(0,len(cols_kl1)-1,14); ws.freeze_panes(3,0)

        # ── Reln Mgr-KCD ───────────────────────────────────
        kcd_l2 = results_dict.get("rm_kcd",[])
        if kcd_l2:
            cols_kl2 = ["Employee ID","Employee Name","L2 ID","L2 Name","L3 ID","L3 Name",
                        "L4 ID","L4 Name","L5 ID","L5 Name","L6  ID","L6 Name",
                        "Joining Date","IIL Vertical Name","Group","Aeging","HC",
                        "Client-A","Client-C","Deal Value","PCDV",
                        "Sent","Recd","Ren %","Sent.1","Recd.1","Ren %.1",
                        "SS+\nMultiplier","Payout\nEligibile","Incentive","Booster\nPayout","Final Incentive",
                        "Gross\nIncentive ","Paid\nIncentive","Balance\nIncentive",
                        "Deal Value","PCDV","Incentive.1","IM Star/Leader\nNew Sale","Gross Incentive","Paid\nIncentive.1","Balance\nIncentive.1",
                        "Deal Value.1","PCDV.1","Incentive.2","IM Star/Leader\nNew Sale.1","Gross Incentive.1","Paid\nIncentive.2","Balance\nIncentive.2",
                        "IM Star Pro/Leader Pro+\nNew Sale","Gross Incentive.2","Paid\nIncentive.3","Balance\nIncentive.3"]
            rows_kl2=[]
            for r in kcd_l2:
                tgt = r.get("target_pcdv",0)
                rows_kl2.append({
                    "Employee ID":r["eid"],"Employee Name":r["name"],
                    "L2 ID":r.get("L2 ID",""),"L2 Name":r.get("L2 Name",""),
                    "L3 ID":r.get("L3 ID",""),"L3 Name":r.get("L3 Name",""),
                    "L4 ID":r.get("L4 ID",""),"L4 Name":r.get("L4 Name",""),
                    "L5 ID":r.get("L5 ID",""),"L5 Name":r.get("L5 Name",""),
                    "L6  ID":r.get("L6 ID",""),"L6 Name":r.get("L6 Name",""),
                    "Joining Date":str(r.get("Joining Date",""))[:10],
                    "IIL Vertical Name":"Tele Annual KCD","Group":r.get("Group",""),
                    "Aeging":r.get("Ageing",0),"Vintage":r.get("Vintage",""),
                    "HC":r.get("HC",0),
                    "Client-A":r.get("Client-A_Agg",r.get("Client-A",0)),
                    "Client-C":r.get("Client-C_Agg",r.get("Client-C",0)),
                    "Deal Value":round(r.get("deal_val",0),2),"PCDV":round(r.get("pcdv",0),2),
                    "Sent":r.get("cmr_sent",0),"Recd":r.get("cmr_recd",0),
                    "Ren %":round(r.get("cmr_pct",0)/100,4),
                    "Sent.1":r.get("ss_sent",0),"Recd.1":r.get("ss_recd",0),
                    "Ren %.1":round(r.get("ss_pct",0)/100,4),
                    "SS+\nMultiplier":r.get("ss_mult",1),
                    "Payout\nEligibile":r.get("payout_eligible","No"),
                    "Incentive":r.get("incentive_grid",0),
                    "Booster\nPayout":r.get("booster",0),"Final Incentive":r.get("final_inc",0),
                    "Gross\nIncentive ":r.get("final_inc",0),"Paid\nIncentive":0,"Balance\nIncentive":r.get("final_inc",0),
                    "Deal Value":round(r.get("fnt_dv",{}).get("1_12",0),2),
                    "PCDV":round(r.get("pcdv_1_12",0),2),
                    "Incentive.1":r.get("spot1_inc",0),"IM Star/Leader\nNew Sale":r.get("im_count",0),
                    "Gross Incentive":r.get("spot1_gross",0),"Paid\nIncentive.1":0,"Balance\nIncentive.1":r.get("spot1_gross",0),
                    "Deal Value.1":round(r.get("fnt_dv",{}).get("20_30",0),2),
                    "PCDV.1":round(r.get("pcdv_20_30",0),2),
                    "Incentive.2":r.get("spot2_inc",0),"IM Star/Leader\nNew Sale.1":r.get("im_count",0),
                    "Gross Incentive.1":r.get("spot2_gross",0),"Paid\nIncentive.2":0,"Balance\nIncentive.2":r.get("spot2_gross",0),
                    "IM Star Pro/Leader Pro+\nNew Sale":r.get("im_pro_count",0),
                    "Gross Incentive.2":r.get("im_ss_spot",0),"Paid\nIncentive.3":0,"Balance\nIncentive.3":r.get("im_ss_spot",0),
                })
            df_kl2 = pd.DataFrame(rows_kl2).reindex(columns=cols_kl2, fill_value="")
            df_kl2.to_excel(w, sheet_name="Reln Mgr-KCD", index=False, startrow=2)
            ws = w.sheets["Reln Mgr-KCD"]
            ws.merge_range(0,35,0,41,"Spot 1st to 12th",org)
            ws.merge_range(0,42,0,48,"Spot 20th to 30th",org)
            ws.merge_range(0,49,0,52,"28th to 30th",org)
            for ci,col in enumerate(cols_kl2): ws.write(1,ci,col,org)
            ws.set_column(0,len(cols_kl2)-1,14); ws.freeze_panes(3,0)

        # ── BM-KCD ─────────────────────────────────────────
        bm_kcd_rows = results_dict.get("bm_kcd",[])
        if bm_kcd_rows:
            cols_bk = ["Employee ID","Employee Name","L4 ID","L4 Name","L5 ID","L5 Name",
                       "L6  ID","L6 Name","Joining Date","IIL Vertical Name","Group",
                       "Aeging","HC","Client-A","Client-C",
                       "Collection","Refund","Net Collection","Deal Value","Target","Ach\n%",
                       "Incentive","Sent","Recd","Ren %","Sent.1","Recd.1","Ren %.1",
                       "Total\nIncentive","Gross\nIncentive","Paid\nIncentive","Balance\nIncentive"]
            rows_bk=[]
            for r in bm_kcd_rows:
                rows_bk.append({
                    "Employee ID":r["eid"],"Employee Name":r["name"],
                    "L4 ID":r.get("L4 ID",""),"L4 Name":r.get("L4 Name",""),
                    "L5 ID":r.get("L5 ID",""),"L5 Name":r.get("L5 Name",""),
                    "L6  ID":r.get("L6 ID",""),"L6 Name":r.get("L6 Name",""),
                    "Joining Date":str(r.get("Joining Date",""))[:10],
                    "IIL Vertical Name":"Tele Annual KCD","Group":r.get("Group",""),
                    "Aeging":r.get("Ageing",0),"HC":r.get("HC",0),
                    "Client-A":r.get("Client-A_Agg",r.get("Client-A",0)),
                    "Client-C":r.get("Client-C_Agg",r.get("Client-C",0)),
                    "Collection":round(r.get("gross_coll",0),2),
                    "Refund":round(r.get("refund",0),2),
                    "Net Collection":round(r.get("net_coll",0),2),  # BM-KCD: Rs
                    "Deal Value":round(r.get("deal_val",0),2),
                    "Target":r.get("coll_target") or None,"Ach\n%":round(r.get("ach_pct",0)/100,4),
                    "Incentive":r.get("base_inc",0),
                    "Sent":r.get("cmr_sent",0),"Recd":r.get("cmr_recd",0),
                    "Ren %":round(r.get("cmr_pct",0)/100,4),
                    "Sent.1":r.get("ss_sent",0),"Recd.1":r.get("ss_recd",0),
                    "Ren %.1":round(r.get("ss_pct",0)/100,4),
                    "Total\nIncentive":r.get("total_inc",0),
                    "Gross\nIncentive":r.get("total_inc",0),
                    "Paid\nIncentive":0,"Balance\nIncentive":r.get("total_inc",0),
                })
            df_bk = pd.DataFrame(rows_bk).reindex(columns=cols_bk, fill_value="")
            df_bk.to_excel(w, sheet_name="BM-KCD", index=False, startrow=1)
            ws = w.sheets["BM-KCD"]
            for ci,col in enumerate(cols_bk): ws.write(1,ci,col,org)
            ws.set_column(0,len(cols_bk)-1,14); ws.freeze_panes(2,0)

        # ── BM-KCD Big Ticket ──────────────────────────────
        bt_bm_kcd_rows = results_dict.get("bt_bm_kcd",[])
        if bt_bm_kcd_rows:
            cols_btbk=["Employee ID","Employee Name","L4 ID","L4 Name","L5 ID","L5 Name",
                       "L6  ID","L6 Name","Location","Joining Date","IIL Vertical Name",
                       "Net Collection","AOP","%","Sent","Recd","Ren %",
                       "Total","10L+","8L+","5L+","3L+","Eligible",
                       "Gross\nIncentive","Paid\nIncentive","Balance\nIncentive"]
            rows_btbk=[]
            for r in bt_bm_kcd_rows:
                cnt=r.get("bt_counts",{})
                rows_btbk.append({
                    "Employee ID":r["eid"],"Employee Name":r["name"],
                    "L4 ID":r.get("L4 ID",""),"L4 Name":r.get("L4 Name",""),
                    "L5 ID":r.get("L5 ID",""),"L5 Name":r.get("L5 Name",""),
                    "L6  ID":r.get("L6 ID",""),"L6 Name":r.get("L6 Name",""),
                    "Location":r.get("Location",""),
                    "Joining Date":str(r.get("Joining Date",""))[:10],
                    "IIL Vertical Name":"Tele Annual KCD",
                    "Net Collection":round(r.get("net_coll",0),2),
                    "AOP":r.get("aop",0),"%":round(r.get("aop_pct",0),2),
                    "Sent":r.get("cmr_sent",0),"Recd":r.get("cmr_recd",0),
                    "Ren %":round(r.get("cmr_pct",0)/100,4),
                    "Total":sum(cnt.get(k,0) for k in ["10L+","8L+","5L+","3L+"]),
                    "10L+":cnt.get("10L+",0),"8L+":cnt.get("8L+",0),
                    "5L+":cnt.get("5L+",0),"3L+":cnt.get("3L+",0),
                    "Eligible":r.get("bt_eligible","NO"),
                    "Gross\nIncentive":r.get("bt_inc",0),"Paid\nIncentive":0,"Balance\nIncentive":r.get("bt_inc",0),
                })
            df_btbk=pd.DataFrame(rows_btbk).reindex(columns=cols_btbk, fill_value="")
            df_btbk.to_excel(w,sheet_name="BM-KCD Big Ticket",index=False,startrow=1)
            ws=w.sheets["BM-KCD Big Ticket"]
            for ci,col in enumerate(cols_btbk): ws.write(1,ci,col,org)
            ws.set_column(0,len(cols_btbk)-1,14); ws.freeze_panes(2,0)

        # ── CH-KCD Big Ticket ──────────────────────────────
        bt_ch_kcd_rows = results_dict.get("bt_ch_kcd",[])
        if bt_ch_kcd_rows:
            cols_btck=["Employee ID","Employee Name","L5 ID","L5 Name","L6  ID","L6 Name",
                       "Location","Joining Date","IIL Vertical Name",
                       "Net Collection","AOP","%","Sent","Recd","Ren %",
                       "Total Deal","10L+","8L+","5L+","3L+","Eligible",
                       "Gross\nIncentive","Paid\nIncentive","Balance\nIncentive"]
            rows_btck=[]
            for r in bt_ch_kcd_rows:
                cnt=r.get("bt_counts",{})
                rows_btck.append({
                    "Employee ID":r["eid"],"Employee Name":r["name"],
                    "L5 ID":r.get("L5 ID",""),"L5 Name":r.get("L5 Name",""),
                    "L6  ID":r.get("L6 ID",""),"L6 Name":r.get("L6 Name",""),
                    "Location":r.get("Location",""),
                    "Joining Date":str(r.get("Joining Date",""))[:10],
                    "IIL Vertical Name":"Tele Annual KCD",
                    "Net Collection":round(r.get("net_coll",0),2),
                    "AOP":r.get("aop",0),"%":round(r.get("aop_pct",0),2),
                    "Sent":r.get("cmr_sent",0),"Recd":r.get("cmr_recd",0),
                    "Ren %":round(r.get("cmr_pct",0)/100,4),
                    "Total Deal":sum(cnt.get(k,0) for k in ["10L+","8L+","5L+","3L+"]),
                    "10L+":cnt.get("10L+",0),"8L+":cnt.get("8L+",0),
                    "5L+":cnt.get("5L+",0),"3L+":cnt.get("3L+",0),
                    "Eligible":r.get("bt_eligible","No"),
                    "Gross\nIncentive":r.get("bt_inc",0),"Paid\nIncentive":0,"Balance\nIncentive":r.get("bt_inc",0),
                })
            df_btck=pd.DataFrame(rows_btck).reindex(columns=cols_btck, fill_value="")
            df_btck.to_excel(w,sheet_name="CH-KCD Big Ticket",index=False,startrow=1)
            ws=w.sheets["CH-KCD Big Ticket"]
            for ci,col in enumerate(cols_btck): ws.write(1,ci,col,org)
            ws.set_column(0,len(cols_btck)-1,14); ws.freeze_panes(2,0)

        # ── CH (L4 Collection Heads) ────────────────────────
        _ch_col_names = [
            "Employee ID","Employee Name","L5 ID","L5 Name","L6  ID","L6 Name",
            "Joining Date","IIL Vertical Name","Client-A","Client-C",
            "Collection","Refund","Net Collection","Target",
            "Ach Pct","Sent","Recd","Ren Pct","Sent.1","Recd.1","Ren Pct.1",
            "Incentive","Gross Incentive","Paid Incentive","Balance Incentive"
        ]
        for _chn, _chv in [(" CH","csd"), ("CH","kcd")]:
            _cr = results_dict.get(f"ch_{_chv}", [])
            if not _cr: continue
            _rws = []
            for r in _cr:
                _coll_rs = r.get("gross_coll",0); _ref_rs = r.get("refund",0)
                _net_rs = _coll_rs - _ref_rs
                _tg = r.get("coll_target", 0)
                # CSD CH: Net in Rs; KCD CH: Net in Lakh
                _net_disp = _net_rs if _chv=="csd" else round(_net_rs/1e5, 4)
                _ac = (_net_rs/_tg*100 if _tg>0 else (_net_disp/_tg*100 if _tg>0 else 0))
                _net_f = round(_net_rs/1e5, 6)
                _pcr = round(_net_rs/max(_coll_rs,1), 6) if _coll_rs else 0
                _hc = r.get("HC",0); _l2 = r.get("L2_Count",0)
                if _chv=="csd":
                    _rws.append({
                        "Employee ID":r["eid"], "Employee Name":r.get("name",""),
                        "L5 ID":r.get("L5 ID",""), "L5 Name":r.get("L5 Name",""),
                        "L6  ID":r.get("L6 ID",""), "L6 Name":r.get("L6 Name",""),
                        "HC":_hc, "L2":_l2,
                        "Client-A":r.get("Client-A_Agg",r.get("Client-A",0)),
                        "Client-C":r.get("Client-C_Agg",r.get("Client-C",0)),
                        "Joining Date":str(r.get("Joining Date",""))[:10],
                        "IIL Vertical Name":"Tele Annual CSD",
                        "Collection":round(_coll_rs,2), "Refund":round(_ref_rs,2),
                        "Net Collection":round(_net_rs,2), "Net Collection F":_net_f, "PCR":_pcr,
                        "Collection \nTarget":_tg, "Ach\n%":round(_net_rs/_tg*100/100,4) if _tg else 0,
                        "Sent":r.get("cmr_sent",0), "Recd":r.get("cmr_recd",0),
                        "Ren %":round(r.get("cmr_pct",0)/100,4),
                        "Incentive":r.get("base_inc",0), "Total Inccentive":r.get("total_inc",0),
                        "Gross Inccentive":r.get("total_inc",0),
                        "Paid\nIncentive":0, "Balance\nIncentive":r.get("total_inc",0)})
                else:
                    _rws.append({
                        "Employee ID":r["eid"], "Employee Name":r.get("name",""),
                        "L6  ID":r.get("L6 ID",""), "L6 Name":r.get("L6 Name",""),
                        "Joining Date":str(r.get("Joining Date",""))[:10],
                        "IIL Vertical Name":"Tele Annual KCD",
                        "Client-A":r.get("Client-A_Agg",r.get("Client-A",0)),
                        "Client-C":r.get("Client-C_Agg",r.get("Client-C",0)),
                        "Collection":round(_coll_rs,2), "Refund":round(_ref_rs,2),
                        "Net Collection":_net_disp,
                        "Collection \nTarget":_tg, "Ach\n%":round(_net_disp/_tg*100/100,4) if _tg else 0,
                        "Sent":r.get("cmr_sent",0), "Recd":r.get("cmr_recd",0),
                        "Ren %":round(r.get("cmr_pct",0)/100,4),
                        "Incentive":r.get("base_inc",0), "Gross\nIncentive":r.get("total_inc",0),
                        "Paid\nIncentive":0, "Balance\nIncentive":r.get("total_inc",0)})
            _cf_csd=["Employee ID","Employee Name","L5 ID","L5 Name","L6  ID","L6 Name",
                     "HC","L2","Client-A","Client-C","Joining Date","IIL Vertical Name",
                     "Collection","Refund","Net Collection","Net Collection F","PCR",
                     "Collection \nTarget","Ach\n%","Sent","Recd","Ren %",
                     "Incentive","Total Inccentive","Gross Inccentive","Paid\nIncentive","Balance\nIncentive"]
            _cf_kcd=["Employee ID","Employee Name","L6  ID","L6 Name","Joining Date","IIL Vertical Name",
                     "Client-A","Client-C","Collection","Refund","Net Collection",
                     "Collection \nTarget","Ach\n%","Sent","Recd","Ren %",
                     "Incentive","Gross\nIncentive","Paid\nIncentive","Balance\nIncentive"]
            _cf = _cf_csd if _chv=="csd" else _cf_kcd
            if _rws:
                _df=pd.DataFrame(_rws).reindex(columns=_cf, fill_value="")
                _df.to_excel(w,sheet_name=_chn,index=False,startrow=1)
                _ws=w.sheets[_chn]; _hf=grn if _chv=="csd" else org
                for ci,col in enumerate(_cf): _ws.write(1,ci,col,_hf)
                _ws.set_column(0,len(_cf)-1,14); _ws.freeze_panes(2,0)
        # ── Summary ────────────────────────────────────────
        all_rows = []
        for key, lst in results_dict.items():
            if not isinstance(lst, list): continue
            for r in lst:
                if "eid" not in r: continue
                all_rows.append({
                    "Employee ID":r["eid"],"Employee Name":r.get("name",""),
                    "Vertical":r.get("Vertical",""),"Designation":r.get("Designation",""),
                    "Vintage":r.get("Vintage",""),"Scheme":r.get("scheme",""),
                    "Net Collection":r.get("net_coll",0),"Deal Value":r.get("deal_val",0),
                    "PCDV":round(r.get("pcdv",0),1),
                    "CMR%":r.get("cmr_pct",0),"SS+CMR%":r.get("ss_pct",0),
                    "Base Incentive":r.get("base_inc",r.get("gross_inc",0)),
                    "Spot 2-6":r.get("sp2_6_gross",0),
                    "Spot 7-12":r.get("sp7_12_gross",0),
                    "Spot 20-30":r.get("sp20_30_gross",0),
                    "BT Incentive":r.get("bt_inc",0),
                    "Total Incentive":r.get("total_inc",0),
                })
        if all_rows:
            df_sum = pd.DataFrame(all_rows)
            df_sum.to_excel(w, sheet_name="Summary", index=False, startrow=1)
            ws = w.sheets["Summary"]
            for ci,col in enumerate(df_sum.columns): ws.write(1,ci,col,hdr)
            ws.set_column(0,len(df_sum.columns)-1,16); ws.freeze_panes(2,0)

        # ── Supporting: Receipt / Renewal / Refund ──────────────
        for _sh, _dk in [(" Receipt Data","rec_df_full"),(" Renewal","rnl_df_full"),(" Refund","ref_df_full")]:
            _df_s = results_dict.get(_dk)
            if _df_s is not None and len(_df_s) > 0:
                try:
                    _df_s.to_excel(w, sheet_name=_sh, index=False, startrow=0)
                    _ws2 = w.sheets[_sh]
                    for _ci, _col in enumerate(_df_s.columns):
                        _ws2.write(0, _ci, str(_col), gry)
                    _ws2.set_column(0, len(_df_s.columns)-1, 12)
                except Exception: pass

    return buf.getvalue()


# ──────────────────────────────────────────────────────────
# STREAMLIT UI
# ──────────────────────────────────────────────────────────
st.title("📊 Tele Annual Incentive Calculator")
st.caption("TA CSD & KCD | All Levels: L1 Exec · L2 Rel'n Mgr · L3 BM · L4 CH · Nursery | Separate from regular CSD/KCD")

with st.sidebar:
    st.header("📂 Upload Files")
    st.markdown("**Step 1 — Raw source files:**")
    receipt_f   = st.file_uploader("1. Receipt file",           type=["xlsx","xlsb"])
    refund_f    = st.file_uploader("2. Refund file",            type=["xlsx","xlsb"])
    renewal_f   = st.file_uploader("3. Renewal file",           type=["xlsx","xlsb"])
    struct_f    = st.file_uploader("4. Employee Structure (FSF_TA sheet)", type=["xlsx","xlsb"])
    slab_f      = st.file_uploader("5. TA Slab Config (optional)", type=["xlsx"])
    target_f    = st.file_uploader("6. Milestone Target file (optional)", type=["xlsx","xlsb"],
                                    help="Sheets: L4, L3, L1 Renewal, L2 Target — provides per-employee CMR targets and collection targets")
    st.markdown("---")
    st.markdown("**Step 2 — Enriched receipt (optional re-run):**")
    st.caption("Upload the enriched Receipt Data sheet from a previous output after making corrections.")
    enr_file    = st.file_uploader("📥 Enriched Receipt", type=["xlsx"])
    st.divider()
    st.header("⚙️ Settings")
    st.caption("Slab targets are configurable via the Slab Config file.")
    calc_btn = st.button("▶ Calculate", type="primary", width='stretch')

# Slab config download
st.subheader("Step 0 – Download TA Slab Config")
st.caption("⚠️ The default slab config contains **April 2026** values. "
           "Upload the May 2026 slab config file (from sidebar) to use May scheme values.")
col_slab1, col_slab2 = st.columns(2)
with col_slab1:
    st.download_button("⬇️ Download April 2026 Slab Config",
                       data=make_slab_excel(),
                       file_name="TA_Slab_Config_Apr2026.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       width='stretch')
with col_slab2:
    st.download_button("⬇️ Download May 2026 Slab Config",
                       data=make_may_slab_excel(),
                       file_name="TA_Slab_Config_May2026.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       width='stretch')
    st.caption("May slab numbers = April defaults. Update the Excel with May actuals once you have sir's output.")

# Load slab config
cfg_raw = load_ta_slab_config(slab_f)
S = parse_slabs(cfg_raw)
if slab_f:
    st.success(f"✅ TA Slab Config loaded from: {slab_f.name}")

# Load milestone targets
ta_targets = load_ta_targets(target_f if "target_f" in dir() else None)
if "target_f" in dir() and target_f:
    n1 = len(ta_targets["l1_cmr"]); n2 = len(ta_targets["l2_cmr"])
    n3 = len(ta_targets["l3_coll"]); n4 = len(ta_targets["l4_coll"])
    st.success(f"✅ Target file loaded — L1: {n1} CMR targets | L2: {n2} CMR targets | "
               f"L3: {n3} collection targets | L4: {n4} collection targets")
else:
    st.info("📋 No target file uploaded — CMR targets use slab config default (40%), "
            "collection targets will show 0 (BM/CH achievement will be 0%).", icon="ℹ️")
if not slab_f:
    st.info("📋 Using built-in default slabs. Download and upload a slab config to customise.", icon="ℹ️")

st.subheader("Step 1 – Upload Employee Structure")
if not struct_f:
    st.info("Upload the Employee Structure file with a FSF_TA sheet.", icon="📂")
    st.stop()

# Cache structure loading by file bytes (fast re-runs when only slab/targets change)
_struct_bytes = struct_f.getvalue() if struct_f else b""
_struct_name  = struct_f.name if struct_f else ""

@st.cache_data(show_spinner=False)
def _cached_load_ta_structure(fbytes: bytes, fname: str):
    if not fbytes: return {}
    try:
        return load_ta_structure(_io.BytesIO(fbytes))
    except Exception:
        return {}

struct_map = _cached_load_ta_structure(_struct_bytes, _struct_name)
if not struct_map:
    st.error("Could not load TA structure. Check the FSF_TA sheet in the file.")
    st.stop()

csd_count = sum(1 for v in struct_map.values() if v["Vertical"]=="CSD")
kcd_count = sum(1 for v in struct_map.values() if v["Vertical"]=="KCD")
st.success(f"✅ Loaded {len(struct_map)} TA employees — CSD: {csd_count}, KCD: {kcd_count}")

with st.expander("Employee Structure Preview"):
    preview = pd.DataFrame([
        {"Emp ID":k,"Name":v["Name"],"Vertical":v["Vertical"],"Designation":v["Designation"],
         "Vintage":v["Vintage"],"Ageing":v["Ageing"],"Client-A":v.get("Client-A",0)}
        for k,v in struct_map.items()
    ])
    st.dataframe(preview, width='stretch', hide_index=True)

if not (receipt_f and refund_f and renewal_f):
    st.info("Upload Receipt, Refund, and Renewal files to proceed.", icon="📂")
    st.stop()

# Load files
rec_raw = clean_receipt(_rf(receipt_f))
ref_raw = _rf(refund_f)
rnl_raw = _rf(renewal_f)

# Mode 2: if enriched receipt uploaded, use it (sir has manually corrected columns)
_use_enriched = False
if "enr_file" in dir() and enr_file:
    try:
        _enr_bytes = enr_file.getvalue()
        _enr_xl    = pd.ExcelFile(_io.BytesIO(_enr_bytes))
        _enr_sheets = [s.strip() for s in _enr_xl.sheet_names]

        # Read Receipt Data sheet
        _rec_sh = next((s for s in _enr_xl.sheet_names if "receipt" in s.lower()), _enr_xl.sheet_names[0])
        enr_raw = pd.read_excel(_io.BytesIO(_enr_bytes), sheet_name=_rec_sh)
        enr_raw.columns = [str(c).strip() for c in enr_raw.columns]

        if "Day" in enr_raw.columns and "FNT" in enr_raw.columns and len(enr_raw) > 0:
            # Store detected month info on the enriched df for later comparison
            _enr_dtc = next((c for c in enr_raw.columns if "entry" in c.lower() or "date" in c.lower()), None)
            _enr_top_month = ""
            if _enr_dtc:
                _enr_dates = pd.to_datetime(enr_raw[_enr_dtc], errors="coerce")
                _enr_mv = _enr_dates.dt.to_period("M").dropna().value_counts()
                _enr_top_month = str(_enr_mv.index[0]) if len(_enr_mv) else ""
            rec_raw = enr_raw
            _use_enriched = True
            st.success(f"✅ Mode 2: Using enriched receipt ({len(enr_raw)} rows) "
                       + (f"[{_enr_top_month}]" if _enr_top_month else "")
                       + " — corrections applied.")
            # Also read corrected Refund and Renewal if present
            _ref_sh = next((s for s in _enr_xl.sheet_names if "refund" in s.lower()), None)
            _rnl_sh = next((s for s in _enr_xl.sheet_names if "renewal" in s.lower() or "renew" in s.lower()), None)
            if _ref_sh:
                _enr_ref = pd.read_excel(_io.BytesIO(_enr_bytes), sheet_name=_ref_sh)
                _enr_ref.columns = [str(c).strip() for c in _enr_ref.columns]
                if len(_enr_ref) > 0: ref_raw = _enr_ref
            if _rnl_sh:
                _enr_rnl = pd.read_excel(_io.BytesIO(_enr_bytes), sheet_name=_rnl_sh)
                _enr_rnl.columns = [str(c).strip() for c in _enr_rnl.columns]
                if len(_enr_rnl) > 0: rnl_raw = _enr_rnl
            _use_enriched = True
            st.success(f"✅ Mode 2: Using enriched receipt ({len(enr_raw)} rows) from "
                       f"sheet '{_rec_sh}' — corrections applied.")
        else:
            st.warning(f"Enriched file has {len(enr_raw)} rows / missing Day+FNT columns — "
                       f"using raw receipt instead. Sheets found: {_enr_sheets}")
    except Exception as _enr_e:
        st.warning(f"Could not read enriched file: {_enr_e}")

# Mode 1: enrich raw receipt → output sheet
if not _use_enriched:
    with st.spinner("Enriching receipt/refund/renewal data…"):
        rec_enriched = enrich_receipt_data(rec_raw, struct_map)
        ref_enriched = enrich_refund_data(ref_raw, struct_map)
        rnl_enriched = enrich_renewal_data(rnl_raw, struct_map)

    # ── Prominent download section ─────────────────────────────────────────
    st.subheader("📥 Step 1 Complete — Download Enriched Files")
    st.markdown("Enriched with hierarchy (L2-L6), Day/Week/FNT labels, product flags. **Review, correct any values, then re-upload in Step 2.**")

    @st.cache_data(show_spinner=False)
    def _build_enr_excel(_rec, _ref, _rnl):
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
            _rec.to_excel(w, sheet_name="Receipt Data", index=False)
            _ref.to_excel(w, sheet_name="Refund",       index=False)
            _rnl.to_excel(w, sheet_name="Renewal",      index=False)
        return buf.getvalue()

    _enr_bytes = _build_enr_excel(rec_enriched, ref_enriched, rnl_enriched)
    st.download_button(
        label="⬇️ Download Enriched Receipt / Refund / Renewal",
        data=_enr_bytes,
        file_name="Enriched_Data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        width='stretch',
    )
    st.info("After downloading and correcting, re-upload the file in the sidebar → "
            "**📥 Enriched Receipt (Step 2)**, then click ▶ Calculate.")
else:
    rec_enriched = rec_raw
    ref_enriched = enrich_refund_data(ref_raw, struct_map)
    rnl_enriched = enrich_renewal_data(rnl_raw, struct_map)
    st.success("✅ **Mode 2 active:** Incentive calculation will use your corrected enriched receipt.")

# Normalise EMP ID in renewal
ec_rnl = find_col(rnl_raw, ["EMP ID","Emp ID","EmpID","Employee ID"])
if ec_rnl:
    rnl_raw[ec_rnl] = rnl_raw[ec_rnl].apply(
        lambda x: str(int(float(x))) if str(x).replace(".","").isdigit() else str(x))

months = get_months(rec_raw, rnl_raw)
if not months:
    st.warning("No months detected from files. Check Entry Date / Month columns.", icon="⚠️")
    st.stop()

sel_month = st.sidebar.selectbox("Month to calculate", options=months,
                                  index=len(months)-1, key="ta_month")

# Dynamically set CALC_DATE and FY_START from sel_month (e.g. "May'26" → 2026-05-31)
try:
    import calendar as _cal
    _month_map = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
                  "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
    _sm_parts = sel_month.replace("'","'").replace("'","").split("'") if "'" in sel_month else sel_month.split(" ")
    _mon_str = _sm_parts[0].strip()[:3]
    _yr_str  = _sm_parts[1].strip() if len(_sm_parts)>1 else "26"
    _yr      = 2000 + int(_yr_str) if len(_yr_str)<=2 else int(_yr_str)
    _mon     = _month_map.get(_mon_str, 4)
    CALC_DATE = date(_yr, _mon, _cal.monthrange(_yr, _mon)[1])
except Exception:
    CALC_DATE = date(2026, 5, 31)
FY_START = date(CALC_DATE.year if CALC_DATE.month >= 4 else CALC_DATE.year-1, 4, 1)
st.sidebar.caption(f"CALC_DATE: {CALC_DATE:%d %b %Y} | FY: {FY_START:%b %Y}")

if _use_enriched:
    # Mode 2: check month matches before proceeding
    try:
        _sel_norm2 = str(sel_month).strip().replace("'","-")
        _sel_dt2 = pd.to_datetime(_sel_norm2, format="%b-%y", errors="coerce")
        _sel_period2 = f"{_sel_dt2.year}-{_sel_dt2.month:02d}" if not pd.isna(_sel_dt2) else ""
        if _enr_top_month and _sel_period2 and _enr_top_month != _sel_period2:
            st.error(
                f"🚨 **Month mismatch!** Enriched file is **{_enr_top_month}** "
                f"but calculation month is **{sel_month}**. "
                f"Download a fresh enriched file from a {sel_month} session.",
                icon="🚨"
            )
            st.stop()
    except Exception: pass
    rec = rec_raw    # enriched receipt used as-is
    ref = ref_raw    # enriched refund used as-is
    _, _, rnl = filter_month(rec_raw, ref_raw, rnl_raw, sel_month)  # still filter renewal
else:
    rec, ref, rnl = filter_month(rec_raw, ref_raw, rnl_raw, sel_month)
st.info(f"📅 {sel_month} | Receipt: {len(rec)} rows | Refund: {len(ref)} rows | Renewal: {len(rnl) if rnl is not None else 0} rows")

st.subheader("Step 2 – Calculate")

# Detect L2 name column in renewal (for L2+ CMR lookup)
l2_rnl_col = find_col(rnl, ["L2","L2 Name","L2Name","RM Name"]) if rnl is not None else None

# Store results across rerenders
if "ta_results" not in st.session_state:
    st.session_state["ta_results"] = None
if "ta_sel_month" not in st.session_state:
    st.session_state["ta_sel_month"] = None

if calc_btn:
    results = {
        "nursery":[],"exec_csd":[],"rm_csd":[],
        "bm_csd":[],"bt_bm_csd":[],"bt_ch_csd":[],"ch_csd":[],
        "exec_kcd":[],"rm_kcd":[],
        "bm_kcd":[],"bt_bm_kcd":[],"bt_ch_kcd":[],"ch_kcd":[],
    }

    # Pre-build ALL indices (receipt + renewal) — runs once
    with st.spinner("Building fast lookup indices…"):
        build_receipt_indices(rec, ref)   # O(N) once → O(1) per employee
        _rnl_idx = build_renewal_index(rnl)
        _rnl_full_idx = build_renewal_index(rnl_raw)

        # Pre-build simple EID index for fast O(1) CMR lookups
        _rnl_eid_idx = build_rnl_eid_idx(rnl)
        _sc_rnl = find_col(rnl, ["Status","STATUS"]) if rnl is not None else None
        _pc_rnl = find_col(rnl, ["WS/MDC Main","Product","Service"]) if rnl is not None else None

        # Previous month data — computed ONCE here, not per employee
        _prev_sel = get_prev_month_str(sel_month, months)
        if _prev_sel:
            _, _, _rnl_prev = filter_month(rec_raw, ref_raw, rnl_raw, _prev_sel)
            _rnl_prev_eid_idx = build_rnl_eid_idx(_rnl_prev)
            _sc_prev = find_col(_rnl_prev, ["Status","STATUS"]) if _rnl_prev is not None else None
            _pc_prev = find_col(_rnl_prev, ["WS/MDC Main","Product","Service"]) if _rnl_prev is not None else None
        else:
            _rnl_prev = None
            _rnl_prev_eid_idx = {}
            _sc_prev = _pc_prev = None
        _nxt_sel = _prev_sel  # keep for compatibility

    prog = st.progress(0, "Calculating…")
    eids = list(struct_map.keys())

    for i, eid in enumerate(eids):
        emp = struct_map[eid]
        vert   = emp["Vertical"]
        desig  = emp["Designation"]
        ageing = emp["Ageing"]
        name   = emp["Name"]
        ca     = max(float(emp.get("Client-A_Agg", emp.get("Client-A",1)) or 1), 1)
        is_l2  = desig in ("L2","L3","L4","L5")
        is_25cr= "25" in str(emp.get("Group",""))

        # Get L1 subordinates for this L2 RM
        l1_eids = emp.get("l1_eids", []) if desig == "L2" else []

        # CMR (current month) — O(1) via pre-built EID index
        team_eids = emp.get("all_l1_eids", []) if desig in ("L3","L4","L5") else []
        if desig == "L2" and l1_eids:
            cmr = calc_cmr_fast_eids_idx(_rnl_eid_idx, l1_eids, _sc_rnl, _pc_rnl)
        elif desig in ("L3","L4","L5") and team_eids:
            cmr = calc_cmr_fast_eids_idx(_rnl_eid_idx, team_eids, _sc_rnl, _pc_rnl)
        elif is_l2 and l2_rnl_col and name:
            cmr = calc_cmr_by_name(rnl, l2_rnl_col, name)
            if cmr["sent"] == 0:
                cmr = calc_cmr_fast_eid(_rnl_eid_idx, eid, _sc_rnl, _pc_rnl)
        else:
            cmr = calc_cmr_fast_eid(_rnl_eid_idx, eid, _sc_rnl, _pc_rnl)

        # CMR+1 — previous month, pre-computed once before loop
        if _prev_sel:
            if desig == "L2" and l1_eids:
                cmr_prev = calc_cmr_fast_eids_idx(_rnl_prev_eid_idx, l1_eids, _sc_prev, _pc_prev)
            elif desig in ("L3","L4","L5") and team_eids:
                cmr_prev = calc_cmr_fast_eids_idx(_rnl_prev_eid_idx, team_eids, _sc_prev, _pc_prev)
            elif is_l2 and l2_rnl_col and name:
                cmr_prev = calc_cmr_by_name(_rnl_prev, l2_rnl_col, name)
                if cmr_prev["sent"] == 0:
                    cmr_prev = calc_cmr_fast_eid(_rnl_prev_eid_idx, eid, _sc_prev, _pc_prev)
            else:
                cmr_prev = calc_cmr_fast_eid(_rnl_prev_eid_idx, eid, _sc_prev, _pc_prev)
        else:
            cmr_prev = {"sent":0,"recd":0,"pct":0.0,"ss_sent":0,"ss_recd":0,"ss_pct":0.0}

        # Per-employee targets from target file
        # L1 CMR renewal target
        if desig == "L1":
            per_emp_cmr_tgt = ta_targets["l1_cmr"].get(eid, S["csd_cmr_tgt"])
        elif desig == "L2":
            per_emp_cmr_tgt = ta_targets["l2_cmr"].get(eid, S["csd_cmr_tgt"])
        else:
            per_emp_cmr_tgt = S["csd_cmr_tgt"]

        # L3/L4 collection targets (looked up by name)
        emp_l3_name = emp.get("L3 Name","").lower()
        emp_l4_name = emp.get("L4 Name","").lower()
        emp_own_name = emp.get("Name","").lower()
        if desig == "L3":
            per_emp_coll_tgt = (ta_targets["l3_coll"].get(eid, 0) or
                                ta_targets["l3_coll"].get(emp_own_name, 0))
        elif desig == "L4":
            per_emp_coll_tgt = (ta_targets["l4_coll"].get(eid, 0) or
                                ta_targets["l4_coll"].get(emp_own_name, 0) or
                                ta_targets["l4_coll"].get(emp_own_name.lower(), 0))
        else:
            per_emp_coll_tgt = 0

        # Inject into emp so calc_employee can use them
        emp = dict(emp)
        emp["CMR_Target_Pct"] = per_emp_cmr_tgt
        if per_emp_coll_tgt > 0:
            emp["Coll_Target"] = per_emp_coll_tgt

        # Recalculate vintage
        _age = emp.get("Ageing", 0)
        if vert == "KCD":
            if _age <= 270:   emp["Vintage"] = "0-270"
            elif _age <= 730: emp["Vintage"] = "270-2Yr"
            else:             emp["Vintage"] = "2Yr+"
        else:
            if _age <= 90:    emp["Vintage"] = "0-90"
            elif _age <= 270: emp["Vintage"] = "90-270"
            else:             emp["Vintage"] = "270+"
        # Group from Department
        _dept = emp.get("Department","")
        if vert == "KCD":
            emp["Group"] = "KCD-25cr" if any(k in str(_dept).upper() for k in ["ILP","IVE"]) else "KCD"
        elif not emp.get("Group",""):
            emp["Group"] = "Tele-A CSD"

        # Sir uses Client-C for PCDV (aggregated from L1s for L2 RMs)
        cc_agg = emp.get("Client-C_Agg", emp.get("Client-C", ca))
        cc = max(float(cc_agg or ca), 1)

        # Receipt data
        _receipt_eids = team_eids if (team_eids and desig in ("L3","L4","L5")) else l1_eids
        data = get_emp_data(rec, ref, eid, desig=desig, emp_name=name, client_a=cc, l1_eids=_receipt_eids)

        # Calculate incentive
        inc = calc_employee(emp, data, cmr, S, is_25cr=is_25cr)

        row = {
            "eid":eid, "name":name, "scheme":inc.get("scheme",""),
            "Vertical":vert,"Designation":desig,"Vintage":emp["Vintage"],
            "cmr_target_pct": per_emp_cmr_tgt,
            "Ageing":ageing,"Joining Date":emp.get("Joining Date",""),
            "cmr1_sent":cmr_prev["sent"],"cmr1_recd":cmr_prev["recd"],"cmr1_pct":cmr_prev["pct"],
            "lt90_count":emp.get("lt90_count",0),
            "l1_eids":l1_eids,
            "Client-A":emp.get("Client-A",0),"Client-C":emp.get("Client-C",0),
            "Client-A_Agg":emp.get("Client-A_Agg",emp.get("Client-A",0)),
            "HC":emp.get("HC",0),"Group":emp.get("Group",""),
            "Location":emp.get("Location","Tele Annual"),
            **{k:v for k,v in emp.items()
                    if "ID" in k or "Name" in k
                    or k in ("MDC","Star","Leader","WS-M","WS-A","IVE",
                             "MDC.1","Star.1","Leader.1","WS-M.1","WS-A.1","IVE.1",
                             "Email Id","Location","Department","Client-A","Client-C",
                             "Client-A_Agg","Client-C_Agg","HC","Ageing",
                             "Joining Date","Vintage","Group","Remarks",
                             "Client_Status_Eligible","FY_Ageing","lt90_count",
                             "l1_eids","all_l1_eids","L2_Count","_nurs_bucket")},
            **inc,
            "fnt_dv":data.get("fnt_dv",{}),
            "pcdv_1_12":data.get("pcdv_1_12",0),"pcdv_20_30":data.get("pcdv_20_30",0),
            "im_count":data.get("im_count",0),"im_pro_count":data.get("im_pro_count",0),
            "gross_coll":data.get("gross",0),"refund":data.get("refund",0),
            "deal_val":data.get("deal_val",0),"net_coll":data.get("net_coll",0),
            "pcdv":data.get("pcdv",0),
        }

        # Big Ticket
        row["bt_inc"] = 0; row["bt_counts"] = {}
        row["aop"] = 0; row["aop_pct"] = 0; row["bt_eligible"] = "NO"
        row["csd_cmr_tgt"] = S["csd_cmr_tgt"]

        if vert == "CSD":
            fy_age = emp.get("FY_Ageing", ageing)
            is_nursery = desig == "L1" and (fy_age <= 90 or emp.get("_nurs_bucket","60-90") == "60D")
            if is_nursery:
                # PDF rule: pay under whichever scheme gives higher earnings
                # Compute both nursery and exec incentive
                vl = emp.get("_nurs_bucket","60-90")
                emp["vintage_label"] = vl
                row["vintage_label"] = vl
                # PDF: employee is eligible for BOTH nursery and base scheme
                # Both get computed; payment from whichever is higher
                results["nursery"].append(row)
                # Compute BASE-SCHEME incentive for exec sheet (must bypass nursery branch)
                _emp_exec = dict(emp)
                _emp_exec["vintage_label"] = ""
                _emp_exec["FY_Ageing"] = 999       # force skip nursery eligibility
                _emp_exec["_nurs_bucket"] = "none" # not 60D
                _inc_exec = calc_employee(_emp_exec, data, cmr, S, is_25cr=is_25cr)
                _row_exec = {**{k:v for k,v in emp.items()
                                if "ID" in k or "Name" in k
                                or k in ("MDC","Star","Leader","WS-M","WS-A","IVE",
                                         "MDC.1","Star.1","Leader.1","WS-M.1","WS-A.1","IVE.1",
                                         "Email Id","Location","Department","Client-A","Client-C",
                                         "Client-A_Agg","Client-C_Agg","HC","Ageing",
                                         "Joining Date","Vintage","Group","Remarks",
                                         "Client_Status_Eligible","FY_Ageing","lt90_count",
                                         "l1_eids","all_l1_eids","L2_Count","_nurs_bucket")},
                             **_inc_exec,
                             "fnt_dv":data.get("fnt_dv",{}),
                             "spot_txn":data.get("spot_txn",{}),
                             "pcdv":data.get("pcdv",0),
                             "deal_val":data.get("deal_val",0),
                             "gross_coll":data.get("gross",0),
                             "refund":data.get("refund",0),
                             "net_coll":data.get("gross",0)-data.get("refund",0),
                             "cmr_sent":cmr.get("sent",0),"cmr_recd":cmr.get("recd",0),
                             "cmr_pct":cmr.get("pct",0),
                             "cmr_target_pct":per_emp_cmr_tgt,
                             "ss_sent":cmr_prev.get("sent",0),"ss_recd":cmr_prev.get("recd",0),
                             "ss_pct":cmr_prev.get("pct",0),
                             "eid":eid,"name":name}
                results["exec_csd"].append(_row_exec)
            elif desig == "L1":
                results["exec_csd"].append(row)
            elif desig == "L2":
                results["rm_csd"].append(row)
            elif desig == "L3":
                bt_inc, bt_counts = calc_big_ticket(rec, eid, desig, vert,
                                                     S["bt_bm_csd"], is_l2=False)
                row.update({"bt_inc":bt_inc,"bt_counts":bt_counts,
                             "bt_eligible":"Yes" if bt_inc>0 else "No"})
                results["bm_csd"].append(row)
                results["bt_bm_csd"].append(row)
            elif desig == "L4":
                bt_inc,bt_counts=calc_big_ticket(rec,eid,desig,vert,S["bt_ch_csd"],is_l2=True)
                row.update({"bt_inc":bt_inc,"bt_counts":bt_counts,"bt_eligible":"Yes" if bt_inc>0 else "No","deal_target":0,"net_coll_cr":row.get("net_coll",0)/1e7})
                results["bt_ch_csd"].append(row); results["ch_csd"].append(row)
            else:
                results["ch_csd"].append(row)

        elif vert == "KCD":
            fy_age = emp.get("FY_Ageing", emp.get("Ageing", 999))
            is_nursery = (desig == "L1" and
                         (fy_age <= 90 or emp.get("_nurs_bucket","60-90") == "60D"))
            if is_nursery:
                vl = emp.get("_nurs_bucket","60-90")
                emp["vintage_label"] = vl
                row["vintage_label"] = vl
                # PDF: eligible for both nursery and base scheme
                results["nursery"].append(row)
                _emp_exec_k = dict(emp)
                _emp_exec_k["vintage_label"] = ""
                _emp_exec_k["FY_Ageing"] = 999
                _emp_exec_k["_nurs_bucket"] = "none"
                _inc_exec_k = calc_employee(_emp_exec_k, data, cmr, S, is_25cr=is_25cr)
                _row_exec_k = {**{k:v for k,v in emp.items()
                                  if "ID" in k or "Name" in k
                                  or k in ("MDC","Star","Leader","WS-M","WS-A","IVE",
                                           "MDC.1","Star.1","Leader.1","WS-M.1","WS-A.1","IVE.1",
                                           "Email Id","Location","Department","Client-A","Client-C",
                                           "Client-A_Agg","Client-C_Agg","HC","Ageing",
                                           "Joining Date","Vintage","Group","Remarks",
                                           "l1_eids","all_l1_eids","L2_Count","_nurs_bucket")},
                               **_inc_exec_k,
                               "fnt_dv":data.get("fnt_dv",{}),
                               "pcdv":data.get("pcdv",0),"deal_val":data.get("deal_val",0),
                               "gross_coll":data.get("gross",0),"refund":data.get("refund",0),
                               "cmr_sent":cmr.get("sent",0),"cmr_recd":cmr.get("recd",0),
                               "cmr_pct":cmr.get("pct",0),"cmr_target_pct":per_emp_cmr_tgt,
                               "ss_sent":cmr_prev.get("sent",0),"ss_recd":cmr_prev.get("recd",0),
                               "ss_pct":cmr_prev.get("pct",0),
                               "eid":eid,"name":name}
                results["exec_kcd"].append(_row_exec_k)
            elif desig == "L1":
                results["exec_kcd"].append(row)
            elif desig == "L2":
                results["rm_kcd"].append(row)
            elif desig == "L3":
                bt_inc, bt_counts = calc_big_ticket(rec, eid, desig, vert,
                                                     S["bt_bm_kcd"], is_l2=False)
                row.update({"bt_inc":bt_inc,"bt_counts":bt_counts,
                             "bt_eligible":"Yes" if bt_inc>0 else "NO"})
                results["bm_kcd"].append(row)
                results["bt_bm_kcd"].append(row)
            elif desig == "L4":
                bt_inc,bt_counts=calc_big_ticket(rec,eid,desig,vert,S["bt_ch_kcd"],is_l2=True)
                row.update({"bt_inc":bt_inc,"bt_counts":bt_counts,"bt_eligible":"Yes" if bt_inc>0 else "No","net_coll_cr":row.get("net_coll",0)/1e7})
                results["bt_ch_kcd"].append(row); results["ch_kcd"].append(row)
            else:
                results["ch_kcd"].append(row)

        if i % 10 == 0 or i == len(eids)-1:
            prog.progress((i+1)/len(eids), f"Processing {i+1}/{len(eids)}…")

    prog.empty()

    # Summary metrics
    all_rows_flat = []
    for lst in results.values():
        if isinstance(lst, list): all_rows_flat.extend(lst)

    if all_rows_flat:
        total_emp = len(all_rows_flat)
        total_payout = sum(r.get("total_inc",0) for r in all_rows_flat)
        total_bt     = sum(r.get("bt_inc",0)    for r in all_rows_flat)
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Employees", total_emp)
        m2.metric("Total Incentive", f"₹{total_payout:,.0f}")
        m3.metric("Total Big Ticket", f"₹{total_bt:,.0f}")
        m4.metric("Avg Incentive", f"₹{total_payout/max(total_emp,1):,.0f}")

    # Preview
    if all_rows_flat:
        st.markdown("#### Employee-wise Preview")
        preview_df = pd.DataFrame([{
            "Emp ID":r.get("eid",""), "Name":r.get("name",""),
            "Vert":r.get("Vertical",""), "Level":r.get("Designation",""),
            "Vintage":r.get("Vintage",""), "Scheme":r.get("scheme",""),
            "PCDV":round(r.get("pcdv",0),0),"CMR%":r.get("cmr_pct",0),
            "Base Inc":r.get("base_inc",r.get("gross_inc",0)),
            "Spot":r.get("sp2_6_gross",0)+r.get("sp7_12_gross",0)+r.get("sp20_30_gross",0),
            "BT Inc":r.get("bt_inc",0),"Total":r.get("total_inc",0),
        } for r in all_rows_flat])
        st.dataframe(preview_df, width='stretch', hide_index=True)

    # Export
    results["rec_df_full"]=rec_enriched; results["rnl_df_full"]=rnl_enriched; results["ref_df_full"]=ref_enriched
    with st.spinner("Building Excel output…"):
        out_bytes = build_excel_output(results, sel_month)
    st.session_state["ta_results"]    = out_bytes
    st.session_state["ta_sel_month"]  = sel_month
    st.success("✅ Calculation complete — download below.")
    st.download_button(
        "⬇️ Download Full TA Incentive Report (Excel)",
        data=out_bytes,
        file_name=f"TA_Incentives_{sel_month.replace('-','_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    with st.expander(f"Zero Incentive Employees ({sum(1 for r in all_rows_flat if r.get('total_inc',0)==0)})"):
        zero = [r for r in all_rows_flat if r.get("total_inc",0)==0]
        if zero:
            st.dataframe(pd.DataFrame([{
                "Emp ID":r["eid"],"Name":r.get("name",""),"Level":r.get("Designation",""),
                "PCDV":round(r.get("pcdv",0),0),"CMR%":r.get("cmr_pct",0),"Scheme":r.get("scheme","")
            } for r in zero]), width='stretch', hide_index=True)
        else:
            st.success("All employees earned an incentive! 🎉")
# ═══════════════════════════════════════════════════════════════════════
# PERSISTENT RESULTS (visible after rerun, survives widget interaction)
# ═══════════════════════════════════════════════════════════════════════
if st.session_state.get("ta_results"):
    _rb = st.session_state["ta_results"]
    _rm = str(st.session_state.get("ta_sel_month","")).replace("'","").replace(" ","_")
    st.divider()
    st.subheader("📊 Output Ready")
    _dc1, _dc2 = st.columns([3,1])
    with _dc1:
        st.download_button(
            label="⬇️ Download TA Incentive Output",
            data=_rb,
            file_name=f"TA_Incentives_{_rm}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width='stretch', type="primary", key="dl_persistent",
        )
    with _dc2:
        if st.button("🗑️ Clear", width='stretch', key="clear_results"):
            st.session_state["ta_results"] = None
            st.rerun()
